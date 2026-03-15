"""
agent_loop_runner.py — Server-side Agent Loop with Shared Workspace

Inspired by OpenCloud's postprocessing pipeline (event-driven state machine)
but with AI brains: LLM generates code per step, executes in sandbox,
passes DataFrames between steps via in-memory workspace (no JSON round-trips).

Architecture difference from the JS-side agent loop:
  JS loop:  Frontend → POST /execute-tool (step 1) → JSON artifacts back
            Frontend → POST /execute-tool (step 2, with prior artifacts as JSON) → ...
            Problem: 10K rows × N steps = massive JSON serialization overhead

  Python loop:  POST /agent/run → Python runs ALL steps server-side
                Step 1 output → workspace["cleaned_data"] = pd.DataFrame (in memory)
                Step 2 reads → workspace["cleaned_data"] directly (zero-copy)
                SSE streams progress to frontend in real-time

Flow:
  1. Frontend sends task plan (steps + input_data) to POST /agent/run
  2. Python loads sheets into workspace as DataFrames
  3. For each step:
     a. Build prompt with workspace schema (column names, dtypes, row counts)
     b. LLM generates Python code
     c. Execute code with workspace access (read/write DataFrames)
     d. Validate output (non-empty artifacts)
     e. Store artifacts in workspace for next step
     f. Publish SSE event with code, result, artifact summary
  4. Return final results + all artifacts
"""

from __future__ import annotations

import asyncio
import base64
import io
import json
import logging
import os
import subprocess
import time
import traceback
import re
from contextlib import redirect_stdout, redirect_stderr
from typing import Any, Dict, List, Optional

import pandas as pd
import numpy as np
from fastapi import APIRouter, BackgroundTasks
from pydantic import BaseModel, Field

logger = logging.getLogger("agent_loop_runner")

agent_loop_router = APIRouter(prefix="/agent", tags=["agent-loop"])

# Import from existing modules
from ml.api.tool_executor import (
    LLMConfig, _pick_provider, _call_llm, _default_model,
    _extract_code_from_llm, _validate_code, _restricted_import,
    _sanitize_result, _ALLOWED_MODULES,
)
from ml.api.agent_sse_router import _get_or_create_channel, _channel_meta

# ---------------------------------------------------------------------------
# Workspace: shared in-memory data store between steps
# ---------------------------------------------------------------------------

class Workspace:
    """
    In-memory data workspace for multi-step execution.
    Steps read/write DataFrames here — zero serialization between steps.
    """

    def __init__(self, input_data: dict):
        self.dataframes: Dict[str, pd.DataFrame] = {}
        self.scalars: Dict[str, Any] = {}
        self.step_results: Dict[str, dict] = {}
        self.artifacts: Dict[str, list] = {}  # step_name → [artifact_dicts]

        # Load input sheets as DataFrames
        sheets = input_data.get("sheets", {})
        for sheet_name, rows in sheets.items():
            if isinstance(rows, list) and len(rows) > 0:
                try:
                    self.dataframes[sheet_name] = pd.DataFrame(rows)
                    logger.info(f"[workspace] Loaded sheet '{sheet_name}': {len(rows)} rows, {len(self.dataframes[sheet_name].columns)} cols")
                except Exception as e:
                    logger.warning(f"[workspace] Failed to load sheet '{sheet_name}': {e}")

    def get_schema_summary(self) -> str:
        """Produce a concise schema summary for LLM prompts."""
        parts = []
        for name, df in self.dataframes.items():
            parts.append(f"### ws['{name}']: {len(df)} rows × {len(df.columns)} cols")
            for col in df.columns[:30]:
                dtype = str(df[col].dtype)
                sample = str(df[col].dropna().iloc[0])[:60] if len(df[col].dropna()) > 0 else "N/A"
                parts.append(f"  - `{col}` ({dtype}): e.g. {sample}")
        for name, val in self.scalars.items():
            parts.append(f"### ws.scalars['{name}']: {type(val).__name__} = {str(val)[:100]}")
        return "\n".join(parts)

    def get_prior_artifacts_summary(self, up_to_step: int) -> str:
        """Summarize artifacts from completed steps."""
        parts = []
        for step_name, arts in self.artifacts.items():
            for art in arts:
                label = art.get("label", "?")
                data = art.get("data")
                if isinstance(data, pd.DataFrame):
                    parts.append(f"- {step_name} / '{label}': DataFrame {data.shape[0]} rows × {data.shape[1]} cols")
                    parts.append(f"  Columns: {list(data.columns[:20])}")
                elif isinstance(data, list) and len(data) > 0:
                    parts.append(f"- {step_name} / '{label}': {len(data)} items")
                elif isinstance(data, dict):
                    parts.append(f"- {step_name} / '{label}': dict with keys {list(data.keys())[:10]}")
        return "\n".join(parts) if parts else "(no prior artifacts)"


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class StepDef(BaseModel):
    name: str
    workflow_type: str = "python_tool"
    tool_hint: str = ""
    ai_review: bool = False
    requires_review: bool = False

class AgentRunRequest(BaseModel):
    task_id: str = Field(..., description="Task ID for SSE channel")
    steps: List[StepDef]
    input_data: Dict[str, Any] = Field(default_factory=dict)
    dataset_profile: Optional[Dict[str, Any]] = None
    llm_config: Optional[LLMConfig] = None
    max_retries: int = 2

class StepResultOut(BaseModel):
    step_name: str
    status: str  # succeeded | failed | blocked
    code: Optional[str] = None
    code_language: str = "python"
    stdout: str = ""
    stderr: str = ""
    error: Optional[str] = None
    artifacts: List[dict] = []
    result: Optional[Any] = None
    execution_ms: int = 0
    llm_provider: Optional[str] = None
    llm_model: Optional[str] = None

class AgentRunResponse(BaseModel):
    ok: bool
    task_id: str
    steps_completed: int = 0
    steps_total: int = 0
    step_results: List[StepResultOut] = []
    total_execution_ms: int = 0
    error: Optional[str] = None


# ---------------------------------------------------------------------------
# Server-side code generation prompt (workspace-aware)
# ---------------------------------------------------------------------------

WORKSPACE_CODE_GEN_SYSTEM = """You are a Python data analyst. Generate clean, simple Python code.

Define `def run(ws):` — ws is a workspace with ws.dataframes (dict of pd.DataFrames).

Example:
```
import pandas as pd

def run(ws):
    df = ws.dataframes['Sales']  # already a DataFrame
    print(df.columns.tolist())   # discover columns

    revenue = pd.to_numeric(df['Revenue'], errors='coerce')
    total = revenue.sum()

    result_df = pd.DataFrame([{"KPI": "Total Revenue", "Value": total}])

    return {
        "result": {"total_revenue": total},
        "artifacts": [{"type": "table", "label": "KPI_Summary", "data": result_df}],
        "dataframes": {"KPI_Summary": result_df}
    }
```

Rules:
- ws.dataframes['SheetName'] gives you a pd.DataFrame (NOT raw dicts)
- Use df.columns.tolist() to see available columns — NEVER guess
- Use pd.to_numeric(col, errors='coerce') for numeric ops
- Return at least 1 artifact with a DataFrame
- "dataframes" dict adds DataFrames to workspace for next steps
- Available: pandas, numpy, scipy, json, re, math, datetime, collections, statistics, time
- Keep code simple and straightforward — prefer pandas built-in methods over scipy when possible
- NEVER use pd.to_datetime(infer_datetime_format=True) — removed in pandas 2.0. Just use pd.to_datetime(col, errors='coerce')
- NEVER use df.append() — removed in pandas 2.0. Use pd.concat([df, new_row]) instead

STRING SAFETY — VERY IMPORTANT:
- NEVER embed long text, Chinese/CJK characters, pipe characters (|), or markdown inside Python string literals.
- Use short ASCII-only strings for labels, descriptions, and column names in code.
- Reference non-ASCII column names dynamically: df.columns.tolist(), NOT hardcoded Chinese strings.
- Keep ALL string literals SHORT (< 80 chars) and on a SINGLE LINE.
- For artifact labels use simple English: "Cleaned_Data", "KPI_Summary", "Analysis", etc.

CODE LENGTH — VERY IMPORTANT:
- Keep code under 100 lines. Be concise. Do NOT write separate processing for every sheet.
- Use loops: `for name, df in ws.dataframes.items():` instead of repeating code per sheet.
- Focus on the MAIN data sheets. Skip metadata sheets (Data_Dictionary, Evaluator_Guide, Case_Brief).
- If there are many sheets, identify the primary data sheet(s) first, then process them.

Return ONLY JSON (no markdown): {"code": "import pandas as pd\\ndef run(ws):\\n ...", "description": "..."}"""


def _build_workspace_prompt(step: StepDef, workspace: Workspace, revision_instructions: list = None) -> str:
    """Build prompt for workspace-aware code generation."""
    parts = [f"## Task\n{step.tool_hint}"]

    parts.append("\n## Workspace Data (access via ws.dataframes['name'])")
    parts.append(workspace.get_schema_summary())

    prior = workspace.get_prior_artifacts_summary(0)
    if prior != "(no prior artifacts)":
        parts.append(f"\n## Prior Step Artifacts\n{prior}")

    if revision_instructions:
        parts.append("\n## REVISION REQUIRED — Fix these issues:")
        for inst in revision_instructions:
            parts.append(f"- {inst}")

    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Workspace-aware code execution
# ---------------------------------------------------------------------------

def _execute_workspace_code(code: str, workspace: Workspace) -> dict:
    """
    Execute Python code with workspace access.
    Code defines run(ws) → { result, artifacts, dataframes }
    """
    stdout_buf = io.StringIO()
    stderr_buf = io.StringIO()

    # Build namespace with workspace
    namespace = {
        "__builtins__": {
            "print": print, "len": len, "range": range, "enumerate": enumerate,
            "zip": zip, "map": map, "filter": filter, "sorted": sorted,
            "reversed": reversed, "iter": iter, "next": next, "slice": slice,
            "min": min, "max": max, "sum": sum, "abs": abs, "round": round,
            "pow": pow, "divmod": divmod,
            "int": int, "float": float, "str": str, "bool": bool,
            "list": list, "dict": dict, "set": set, "tuple": tuple,
            "frozenset": frozenset, "bytes": bytes, "bytearray": bytearray,
            "complex": complex, "object": object,
            "type": type, "isinstance": isinstance, "issubclass": issubclass,
            "hasattr": hasattr, "getattr": getattr, "setattr": setattr,
            "delattr": delattr, "callable": callable, "id": id, "hash": hash,
            "dir": dir, "vars": vars, "repr": repr, "format": format,
            "super": super, "property": property,
            "staticmethod": staticmethod, "classmethod": classmethod,
            "any": any, "all": all,
            "chr": chr, "ord": ord, "hex": hex, "bin": bin, "oct": oct, "ascii": ascii,
            "Exception": Exception, "ValueError": ValueError, "TypeError": TypeError,
            "KeyError": KeyError, "IndexError": IndexError,
            "AttributeError": AttributeError, "RuntimeError": RuntimeError,
            "NotImplementedError": NotImplementedError,
            "ZeroDivisionError": ZeroDivisionError, "StopIteration": StopIteration,
            "OverflowError": OverflowError, "ArithmeticError": ArithmeticError,
            "None": None, "True": True, "False": False,
            "__import__": _restricted_import,
            "__name__": "__main__",
            "__build_class__": __builtins__["__build_class__"] if isinstance(__builtins__, dict) else getattr(__builtins__, "__build_class__"),
        },
        "pd": pd, "np": np, "pandas": pd, "numpy": np,
        "json": json, "re": re,
        "math": __import__("math"),
        "datetime": __import__("datetime"),
        "collections": __import__("collections"),
        "statistics": __import__("statistics"),
        "time": __import__("time"),
    }

    start = time.time()

    try:
        with redirect_stdout(stdout_buf), redirect_stderr(stderr_buf):
            exec(code, namespace)

            run_fn = namespace.get("run")
            if not callable(run_fn):
                return {
                    "ok": False,
                    "error": "Code must define `run(ws)` function",
                    "stdout": stdout_buf.getvalue(),
                    "stderr": stderr_buf.getvalue(),
                    "execution_ms": int((time.time() - start) * 1000),
                }

            raw_result = run_fn(workspace)

        execution_ms = int((time.time() - start) * 1000)

        if not isinstance(raw_result, dict):
            return {
                "ok": False,
                "error": f"run(ws) must return a dict, got {type(raw_result).__name__}",
                "stdout": stdout_buf.getvalue(),
                "stderr": stderr_buf.getvalue(),
                "execution_ms": execution_ms,
            }

        # Process artifacts — keep DataFrames as-is in workspace
        artifacts = raw_result.get("artifacts", [])
        serialized_artifacts = []
        for art in artifacts:
            if not isinstance(art, dict):
                continue
            data = art.get("data")
            label = str(art.get("label", "Output"))

            # If artifact data is a DataFrame, store it in workspace too
            if isinstance(data, pd.DataFrame):
                workspace.dataframes[label] = data
                serialized = json.loads(data.head(500).to_json(orient="records", date_format="iso", default_handler=str))
                row_count = len(data)
            elif isinstance(data, pd.Series):
                serialized = json.loads(data.to_json(default_handler=str))
                row_count = len(data)
            elif isinstance(data, np.ndarray):
                serialized = data.tolist()
                row_count = len(data)
            elif isinstance(data, list):
                serialized = data[:500]  # Cap for SSE/JSON transport
                row_count = len(data)
            else:
                serialized = data
                row_count = 1 if data else 0

            serialized_artifacts.append({
                "type": str(art.get("type", "data")),
                "label": label,
                "data": serialized,
                "row_count": row_count,
            })

        # Store new DataFrames from return value
        new_dfs = raw_result.get("dataframes", {})
        for df_name, df_val in new_dfs.items():
            if isinstance(df_val, pd.DataFrame):
                workspace.dataframes[df_name] = df_val

        return {
            "ok": True,
            "result": _sanitize_result(raw_result.get("result")),
            "artifacts": serialized_artifacts,
            "stdout": stdout_buf.getvalue(),
            "stderr": stderr_buf.getvalue(),
            "execution_ms": execution_ms,
        }

    except Exception as e:
        execution_ms = int((time.time() - start) * 1000)
        return {
            "ok": False,
            "error": f"{type(e).__name__}: {str(e)}",
            "stdout": stdout_buf.getvalue(),
            "stderr": stderr_buf.getvalue() + "\n" + traceback.format_exc(),
            "execution_ms": execution_ms,
        }


# ---------------------------------------------------------------------------
# SSE helper: publish step event to channel
# ---------------------------------------------------------------------------

async def _publish_step_event(task_id: str, event_type: str, data: dict):
    """Push event to SSE channel for real-time frontend updates."""
    channel = _get_or_create_channel(task_id)
    event_data = {"event_type": event_type, "timestamp": time.time(), **data}
    try:
        channel.put_nowait(event_data)
        _channel_meta[task_id]["last_event_at"] = time.time()
    except asyncio.QueueFull:
        try:
            channel.get_nowait()
        except asyncio.QueueEmpty:
            pass
        channel.put_nowait(event_data)


# ---------------------------------------------------------------------------
# Core: run all steps server-side
# ---------------------------------------------------------------------------

async def _run_agent_loop(request: AgentRunRequest):
    """
    Execute all steps server-side with shared workspace.
    Publishes SSE events for real-time progress.
    """
    task_id = request.task_id
    total_start = time.time()

    # Initialize workspace from input data
    workspace = Workspace(request.input_data)
    logger.info(f"[agent_loop] Starting task {task_id} with {len(request.steps)} steps, "
                f"workspace has {len(workspace.dataframes)} DataFrames")

    # Resolve LLM provider
    try:
        llm_config = _pick_provider(request.llm_config)
    except ValueError as e:
        await _publish_step_event(task_id, "loop_error", {"error": str(e)})
        return AgentRunResponse(ok=False, task_id=task_id, error=str(e),
                                steps_total=len(request.steps))

    used_provider = llm_config.provider
    used_model = llm_config.model or _default_model(llm_config.provider)

    step_results: List[StepResultOut] = []
    steps_completed = 0

    for i, step in enumerate(request.steps):
        step_start = time.time()
        step_name = step.name

        # Publish step_started
        await _publish_step_event(task_id, "step_started", {
            "step_name": step_name, "step_index": i, "status": "running",
        })

        retry_count = 0
        revision_instructions = []
        step_result = None

        while retry_count <= request.max_retries:
            try:
                # 1. Generate code via LLM
                prompt = _build_workspace_prompt(step, workspace, revision_instructions or None)
                llm_response = await _call_llm(prompt, WORKSPACE_CODE_GEN_SYSTEM, llm_config)
                code = _extract_code_from_llm(llm_response)

                if not code:
                    raise ValueError(f"LLM did not generate valid code. Raw: {llm_response[:300]}")

                logger.info(f"[agent_loop] Step '{step_name}': LLM generated {len(code)} chars via {used_provider}/{used_model}")

                # Publish code_generated event (so UI shows code before execution)
                await _publish_step_event(task_id, "step_event", {
                    "step_name": step_name, "step_index": i,
                    "status": "running",
                    "summary": f"Code generated ({len(code.splitlines())} lines) — now executing...",
                    "code": code,
                    "code_language": "python",
                    "api_call": {
                        "method": "POST", "url": "/agent/run (server-side)",
                        "provider": used_provider, "model": used_model,
                        "duration_ms": int((time.time() - step_start) * 1000),
                        "status": 200,
                    },
                })

                # 2. Validate code safety
                safety_err = _validate_code(code)
                if safety_err:
                    raise ValueError(safety_err)

                # 3. Execute with workspace (in thread pool to not block event loop)
                # Timeout: 120s max per step — prevents infinite loops
                try:
                    exec_result = await asyncio.wait_for(
                        asyncio.to_thread(_execute_workspace_code, code, workspace),
                        timeout=120
                    )
                except asyncio.TimeoutError:
                    raise RuntimeError("Code execution timed out after 120 seconds")

                if not exec_result["ok"]:
                    raise RuntimeError(exec_result.get("error", "Execution failed"))

                artifacts = exec_result.get("artifacts", [])
                if not artifacts:
                    raise RuntimeError("Code executed but produced 0 artifacts")

                # Store artifacts in workspace
                workspace.artifacts[step_name] = artifacts

                # Build step result
                step_result = StepResultOut(
                    step_name=step_name,
                    status="succeeded",
                    code=code,
                    stdout=exec_result.get("stdout", ""),
                    stderr=exec_result.get("stderr", ""),
                    artifacts=artifacts,
                    result=exec_result.get("result"),
                    execution_ms=int((time.time() - step_start) * 1000),
                    llm_provider=used_provider,
                    llm_model=used_model,
                )

                # Publish step_completed with full details
                await _publish_step_event(task_id, "step_completed", {
                    "step_name": step_name,
                    "step_index": i,
                    "status": "succeeded",
                    "code": code,
                    "code_language": "python",
                    "stdout": exec_result.get("stdout", "")[:2000],
                    "stderr": exec_result.get("stderr", "")[:1000],
                    "artifacts": [{
                        "type": a.get("type"), "label": a.get("label"),
                        "row_count": a.get("row_count", 0),
                    } for a in artifacts],
                    "api_call": {
                        "method": "POST", "url": "/agent/run (server-side)",
                        "provider": used_provider, "model": used_model,
                        "duration_ms": int((time.time() - step_start) * 1000),
                        "status": 200,
                    },
                    "summary": f"Step '{step_name}' completed: {len(artifacts)} artifact(s), "
                               f"{sum(a.get('row_count', 0) for a in artifacts)} total rows",
                })

                steps_completed += 1
                break  # Success — move to next step

            except Exception as e:
                retry_count += 1
                error_msg = f"{type(e).__name__}: {str(e)}"
                logger.warning(f"[agent_loop] Step '{step_name}' attempt {retry_count} failed: {error_msg}")

                if retry_count > request.max_retries:
                    # Final failure
                    step_result = StepResultOut(
                        step_name=step_name,
                        status="blocked",
                        code=code if 'code' in dir() else None,
                        error=error_msg,
                        execution_ms=int((time.time() - step_start) * 1000),
                        llm_provider=used_provider,
                        llm_model=used_model,
                    )

                    await _publish_step_event(task_id, "step_failed", {
                        "step_name": step_name, "step_index": i,
                        "status": "blocked", "error": error_msg,
                        "code": code if 'code' in dir() else None,
                    })
                    break

                # Add error to revision instructions for retry
                revision_instructions.append(f"Previous attempt failed: {error_msg}")

                await _publish_step_event(task_id, "step_revision", {
                    "step_name": step_name, "step_index": i,
                    "status": "retrying", "error": error_msg,
                    "retry_count": retry_count,
                    "code": code if 'code' in dir() else None,
                })

                # Backoff before retry
                backoff = min(1000 * (2 ** (retry_count - 1)), 10000) / 1000
                await asyncio.sleep(backoff)

        if step_result:
            step_results.append(step_result)

        # If step blocked, stop the pipeline
        if step_result and step_result.status == "blocked":
            logger.warning(f"[agent_loop] Pipeline halted at step '{step_name}'")
            break

    total_ms = int((time.time() - total_start) * 1000)

    # Publish loop_done
    await _publish_step_event(task_id, "loop_done", {
        "steps_completed": steps_completed,
        "steps_total": len(request.steps),
        "total_execution_ms": total_ms,
    })
    # Terminal signal for SSE stream
    channel = _get_or_create_channel(task_id)
    try:
        channel.put_nowait({"_end": True})
    except asyncio.QueueFull:
        pass

    return AgentRunResponse(
        ok=steps_completed == len(request.steps),
        task_id=task_id,
        steps_completed=steps_completed,
        steps_total=len(request.steps),
        step_results=step_results,
        total_execution_ms=total_ms,
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@agent_loop_router.post("/run", response_model=AgentRunResponse)
async def run_agent_loop(request: AgentRunRequest, background_tasks: BackgroundTasks):
    """
    Run the full agent loop server-side.

    Key advantages over the JS-side loop:
    1. Data stays in memory as DataFrames — no JSON serialization between steps
    2. Steps share a workspace — step 2 reads step 1's output directly
    3. SSE streams real-time progress to the frontend
    4. Server-side retry with revision instructions
    """
    logger.info(f"[agent_loop] Received task {request.task_id} with {len(request.steps)} steps")
    result = await _run_agent_loop(request)
    return result


@agent_loop_router.post("/run-async")
async def run_agent_loop_async(request: AgentRunRequest, background_tasks: BackgroundTasks):
    """
    Run the agent loop in the background. Subscribe to SSE for progress.
    Returns immediately with task_id.
    """
    logger.info(f"[agent_loop] Starting background task {request.task_id}")
    background_tasks.add_task(_run_agent_loop, request)
    return {
        "ok": True,
        "task_id": request.task_id,
        "status": "started",
        "sse_url": f"/sse/agent/{request.task_id}/events",
    }


# ---------------------------------------------------------------------------
# MBR Analysis — convenience wrapper for Excel Add-in
# ---------------------------------------------------------------------------

class MbrAnalysisRequest(BaseModel):
    """Simplified request for MBR analysis from Excel Add-in."""
    task_id: str = Field(default_factory=lambda: f"mbr_{int(time.time() * 1000)}")
    input_data: Dict[str, Any] = Field(..., description="{ sheets: { name: rows[] } }")
    focus: str = Field(default="", description="Optional user focus prompt")
    max_retries: int = 2

MBR_DEFAULT_STEPS = [
    StepDef(
        name="clean_data",
        workflow_type="python_tool",
        tool_hint="Clean and standardize the raw data. Create cleaned_data DataFrame and data_issues_log. Fix dates, numeric types, missing values, duplicates. Output: 1) cleaned DataFrame with consistent types, 2) issues log array [{column, issue, action, rows_affected}].",
    ),
    StepDef(
        name="calculate_kpis",
        workflow_type="python_tool",
        tool_hint="Calculate business KPIs from cleaned data. Compute: total_revenue, units_sold, gross_profit, gross_margin_pct, avg_selling_price, return_rate, discount_rate, unique_products, unique_regions. Output: 1) kpi_summary dict, 2) kpi_by_dimension DataFrame (breakdown by available dimensions). Keep code under 80 lines.",
    ),
    StepDef(
        name="pivot_analysis",
        workflow_type="python_tool",
        tool_hint="Create pivot analysis and management insights from cleaned data and KPIs. Build pivot tables by available dimensions (region, product, time). Generate 3-5 management insights as strings. Output: 1) pivot_tables (list of DataFrames), 2) insights array of strings, 3) analysis_summary DataFrame. Keep code under 80 lines.",
    ),
]


@agent_loop_router.post("/mbr-analysis")
async def run_mbr_analysis(request: MbrAnalysisRequest):
    """
    Convenience endpoint for MBR analysis from the Excel Add-in.
    Wraps /agent/run with pre-configured MBR steps.
    """
    steps = MBR_DEFAULT_STEPS.copy()

    # Append user focus to each step's tool_hint
    if request.focus:
        steps = [
            StepDef(
                name=s.name,
                workflow_type=s.workflow_type,
                tool_hint=f"{s.tool_hint} User focus: {request.focus}",
            )
            for s in steps
        ]

    agent_request = AgentRunRequest(
        task_id=request.task_id,
        steps=steps,
        input_data=request.input_data,
        max_retries=request.max_retries,
    )

    logger.info(f"[mbr] Running MBR analysis for task {request.task_id}")
    result = await _run_agent_loop(agent_request)
    return result


# ---------------------------------------------------------------------------
# Excel Workbook Generation — LLM-driven openpyxl (Opus 4.6)
# ---------------------------------------------------------------------------

EXCEL_CODEGEN_SYSTEM_PROMPT = """You are an expert Excel workbook designer using Python openpyxl.
Given step_results from a data analysis pipeline, generate openpyxl code to build a professional, formatted .xlsx workbook.

CRITICAL RULES:
1. Define a function `build_workbook(step_results, title)` that returns an openpyxl.Workbook object.
2. Available libraries: openpyxl (all submodules), json, datetime, math, re, collections
3. DO NOT import os, sys, subprocess, shutil, or any I/O libraries
4. DO NOT use open(), exec(), eval(), __import__(), compile()
5. DO NOT call wb.save() — the caller handles saving
6. step_results is a list of dicts: [{"step_name": str, "status": str, "artifacts": [{"type": str, "label": str, "data": list|dict}]}]

WORKBOOK DESIGN GUIDELINES:
- Create a Cover sheet with the title, generation timestamp, and table of contents
- Create a KPI Dashboard sheet with key metrics displayed as large-font KPI cards
- Create data sheets for cleaned data, data issues, and analysis results
- Create a Dashboard summary sheet with KPI highlights, insights, and at least one chart (BarChart or LineChart)
- Use professional styling:
  - Title: Font(size=24, bold=True, color="1F4E79")
  - Section headers: Font(size=14, bold=True, color="1F4E79")
  - Table headers: white text on blue background (fgColor="2563EB")
  - Alternating row colors (fgColor="F1F5F9" for even rows)
  - Auto-fit column widths (scan first 50 rows, cap at 40 chars)
  - Number formatting: "#,##0" for integers, "0.0%" for percentages
- Remove the default blank sheet (wb.remove(wb.active))
- Handle missing/empty artifacts gracefully — show "No data available" placeholder
- The workbook should have at least 6 sheets: Cover, KPIs, Cleaned_Data, Data_Issues, Analysis, Dashboard
- Add openpyxl charts (BarChart, LineChart) where numeric pivot data is available
- Keep string literals SHORT and ASCII-only. NO Chinese/CJK characters.

ARTIFACT ACCESS PATTERN:
```python
# Collect all succeeded artifacts
all_artifacts = {}
for sr in step_results:
    if sr.get("status") == "succeeded":
        all_artifacts[sr["step_name"]] = sr.get("artifacts", [])

# Find data by keyword matching on type/label
def find_data(keywords):
    for arts in all_artifacts.values():
        for art in arts:
            t = (art.get("type") or "").lower()
            label = (art.get("label") or "").lower()
            if any(kw in t or kw in label for kw in keywords):
                return art.get("data") or art.get("content")
    return None
```

Return ONLY a JSON object (no markdown, no explanation):
{
  "code": "import openpyxl\\nfrom openpyxl.styles import ...\\n\\ndef build_workbook(step_results, title):\\n    ...",
  "description": "Brief description of workbook structure"
}"""


class GenerateExcelRequest(BaseModel):
    """Generate a formatted .xlsx workbook from agent step artifacts."""
    task_id: str
    step_results: List[dict] = Field(..., description="List of step result dicts with artifacts")
    title: str = ""
    open_file: bool = Field(default=True, description="Open the file in Excel desktop after generating")
    output_dir: str = Field(default="", description="Optional output directory (default: ./output/)")


def _build_excel_codegen_prompt(step_results: list, title: str) -> str:
    """Build prompt for the LLM to generate openpyxl code."""
    parts = [f"## Task\nGenerate a professional Excel workbook titled: \"{title or 'Monthly Business Review'}\""]

    # Show artifact summary so LLM knows what data is available
    parts.append("\n## Available Artifacts from Analysis Steps")
    for sr in step_results:
        step_name = sr.get("step_name", "unknown")
        status = sr.get("status", "unknown")
        artifacts = sr.get("artifacts", [])
        parts.append(f"\n### Step: {step_name} (status: {status})")
        for art in artifacts:
            art_type = art.get("type", "unknown")
            art_label = art.get("label", "unknown")
            data = art.get("data") or art.get("content")
            if isinstance(data, list) and len(data) > 0:
                row_count = len(data)
                if isinstance(data[0], dict):
                    cols = list(data[0].keys())
                    parts.append(f"- **{art_label}** (type: {art_type}): {row_count} rows, columns: {cols[:20]}")
                    # Show 2 sample rows
                    for i, row in enumerate(data[:2]):
                        parts.append(f"  Sample row {i+1}: {json.dumps(row, default=str, ensure_ascii=False)[:500]}")
                else:
                    parts.append(f"- **{art_label}** (type: {art_type}): {row_count} items")
            elif isinstance(data, dict):
                parts.append(f"- **{art_label}** (type: {art_type}): dict with keys {list(data.keys())[:15]}")
            else:
                parts.append(f"- **{art_label}** (type: {art_type}): {type(data).__name__ if data else 'empty'}")

    if not step_results:
        parts.append("No analysis artifacts available. Create a workbook with placeholder sheets.")

    parts.append("\n## Instructions")
    parts.append("Generate openpyxl code that creates a workbook with these artifacts organized into professional sheets.")
    parts.append("The `build_workbook(step_results, title)` function receives the EXACT step_results list shown above.")

    return "\n".join(parts)


def _execute_excel_code(code: str, step_results: list, title: str) -> dict:
    """Execute LLM-generated openpyxl code in a sandbox that allows openpyxl imports."""
    import openpyxl
    import openpyxl.styles
    import openpyxl.utils
    import openpyxl.chart

    stdout_buf = io.StringIO()
    stderr_buf = io.StringIO()

    namespace = {
        "__builtins__": {
            "print": print, "len": len, "range": range, "enumerate": enumerate,
            "zip": zip, "map": map, "filter": filter, "sorted": sorted,
            "reversed": reversed, "iter": iter, "next": next, "slice": slice,
            "min": min, "max": max, "sum": sum, "abs": abs, "round": round,
            "pow": pow, "divmod": divmod,
            "int": int, "float": float, "str": str, "bool": bool,
            "list": list, "dict": dict, "set": set, "tuple": tuple,
            "frozenset": frozenset, "bytes": bytes, "bytearray": bytearray,
            "complex": complex, "object": object,
            "type": type, "isinstance": isinstance, "issubclass": issubclass,
            "hasattr": hasattr, "getattr": getattr, "setattr": setattr,
            "delattr": delattr, "callable": callable, "id": id, "hash": hash,
            "dir": dir, "vars": vars, "repr": repr, "format": format,
            "super": super, "property": property,
            "staticmethod": staticmethod, "classmethod": classmethod,
            "any": any, "all": all,
            "chr": chr, "ord": ord, "hex": hex, "bin": bin, "oct": oct, "ascii": ascii,
            "Exception": Exception, "ValueError": ValueError, "TypeError": TypeError,
            "KeyError": KeyError, "IndexError": IndexError, "AttributeError": AttributeError,
            "RuntimeError": RuntimeError, "NotImplementedError": NotImplementedError,
            "ZeroDivisionError": ZeroDivisionError, "StopIteration": StopIteration,
            "None": None, "True": True, "False": False,
            "__import__": _restricted_import,
            "__name__": "__main__",
            "__build_class__": __builtins__["__build_class__"] if isinstance(__builtins__, dict) else getattr(__builtins__, "__build_class__"),
        },
        "openpyxl": openpyxl,
        "json": json,
        "re": re,
        "math": __import__("math"),
        "datetime": __import__("datetime"),
        "collections": __import__("collections"),
    }

    start = time.time()
    try:
        with redirect_stdout(stdout_buf), redirect_stderr(stderr_buf):
            exec(code, namespace)
            build_fn = namespace.get("build_workbook")
            if not callable(build_fn):
                return {"ok": False, "error": "Code must define `build_workbook(step_results, title)`",
                        "stdout": stdout_buf.getvalue(), "stderr": stderr_buf.getvalue()}
            wb = build_fn(step_results, title)

        if not isinstance(wb, openpyxl.Workbook):
            return {"ok": False, "error": f"build_workbook() must return openpyxl.Workbook, got {type(wb).__name__}",
                    "stdout": stdout_buf.getvalue(), "stderr": stderr_buf.getvalue()}

        return {"ok": True, "workbook": wb, "execution_ms": int((time.time() - start) * 1000),
                "stdout": stdout_buf.getvalue(), "stderr": stderr_buf.getvalue()}

    except Exception as e:
        return {"ok": False, "error": f"{type(e).__name__}: {e}",
                "stdout": stdout_buf.getvalue(),
                "stderr": stderr_buf.getvalue() + "\n" + traceback.format_exc(),
                "execution_ms": int((time.time() - start) * 1000)}


@agent_loop_router.post("/generate-excel")
async def generate_excel_workbook(request: GenerateExcelRequest):
    """
    Generate a formatted Excel workbook from agent step artifacts.

    This is the "Excel as a tool" endpoint — the AI Employee calls this after
    analysis steps complete. Opus 4.6 writes the openpyxl code dynamically
    based on the actual artifacts, then we execute it in a sandbox.

    Flow:
      1. Build prompt with artifact summaries for Opus 4.6
      2. LLM generates tailored openpyxl code
      3. Execute in sandbox → openpyxl.Workbook
      4. Save .xlsx, optionally open in Excel desktop
      5. Return file path + base64 for upload
    """
    from datetime import datetime as dt
    from ml.api.tool_executor import _extract_code_from_llm

    step_results = request.step_results
    output_dir = request.output_dir or os.path.join(os.getcwd(), "output")
    os.makedirs(output_dir, exist_ok=True)

    timestamp = dt.now().strftime("%Y%m%d_%H%M%S")
    filename = f"MBR_{request.task_id}_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)
    title = request.title or f"Monthly Business Review — {dt.now().strftime('%Y-%m')}"

    logger.info(f"[excel-gen] Generating workbook via Opus 4.6: {output_path}")

    # ── Step 1: Call Opus 4.6 to generate openpyxl code ──
    llm_config = LLMConfig(provider="anthropic", model="claude-opus-4-6", temperature=0.1, max_tokens=16384)
    llm_config = _pick_provider(llm_config)

    prompt = _build_excel_codegen_prompt(step_results, title)
    logger.info(f"[excel-gen] Calling LLM ({llm_config.provider}/{llm_config.model or 'default'}) for openpyxl code...")

    try:
        llm_response = await _call_llm(prompt, EXCEL_CODEGEN_SYSTEM_PROMPT, llm_config)
        code = _extract_code_from_llm(llm_response)
        if not code:
            return {"ok": False, "error": "LLM did not return valid code", "llm_response": llm_response[:2000]}
        logger.info(f"[excel-gen] LLM generated {len(code)} chars of openpyxl code")
    except Exception as e:
        logger.error(f"[excel-gen] LLM call failed: {e}", exc_info=True)
        return {"ok": False, "error": f"LLM code generation failed: {e}"}

    # ── Step 2: Execute the generated code in sandbox ──
    exec_result = _execute_excel_code(code, step_results, title)

    if not exec_result["ok"]:
        logger.warning(f"[excel-gen] First attempt failed: {exec_result['error']}")
        logger.warning(f"[excel-gen] stderr: {exec_result.get('stderr', '')[:500]}")

        # Retry: send error back to LLM for self-correction
        fix_prompt = (
            f"The openpyxl code you generated failed with this error:\n\n"
            f"```\n{exec_result['error']}\n```\n\n"
            f"stderr:\n```\n{exec_result.get('stderr', '')[:1000]}\n```\n\n"
            f"Here was your code:\n```python\n{code[:3000]}\n```\n\n"
            f"Fix the code. Return ONLY a JSON object with the corrected code."
        )
        try:
            llm_response2 = await _call_llm(fix_prompt, EXCEL_CODEGEN_SYSTEM_PROMPT, llm_config)
            code2 = _extract_code_from_llm(llm_response2)
            if code2:
                logger.info(f"[excel-gen] Retry: LLM generated {len(code2)} chars of fixed code")
                exec_result = _execute_excel_code(code2, step_results, title)
                code = code2
        except Exception as e2:
            logger.error(f"[excel-gen] Retry LLM call failed: {e2}")

    if not exec_result["ok"]:
        return {"ok": False, "error": exec_result["error"],
                "generated_code": code[:5000],
                "stdout": exec_result.get("stdout", ""),
                "stderr": exec_result.get("stderr", "")}

    # ── Step 3: Save workbook + open + base64 ──
    wb = exec_result["workbook"]
    try:
        wb.save(output_path)
        file_size = os.path.getsize(output_path)
        logger.info(f"[excel-gen] Saved: {output_path} ({file_size} bytes, {len(wb.sheetnames)} sheets)")

        with open(output_path, "rb") as f:
            file_bytes = f.read()
        file_base64 = base64.b64encode(file_bytes).decode("ascii")

        if request.open_file:
            try:
                subprocess.Popen(["open", "-a", "Microsoft Excel", output_path])
                logger.info(f"[excel-gen] Opened in Excel: {output_path}")
            except Exception as e:
                logger.warning(f"[excel-gen] Could not open Excel: {e}")

        return {
            "ok": True,
            "file_path": output_path,
            "filename": filename,
            "file_size": file_size,
            "sheets": wb.sheetnames,
            "content_base64": file_base64,
            "llm_model": llm_config.model or _default_model(llm_config.provider),
            "code_length": len(code),
            "execution_ms": exec_result.get("execution_ms", 0),
        }

    except Exception as e:
        logger.error(f"[excel-gen] Save failed: {e}", exc_info=True)
        return {"ok": False, "error": str(e)}
