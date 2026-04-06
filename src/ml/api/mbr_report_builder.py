"""
mbr_report_builder.py — MBR Excel Report Generator

Architecture: LLM plans the report → openpyxl renders it

Flow:
  1. LLM receives all artifacts + narrative → outputs JSON report_plan
     (which artifacts go on which sheets, chart types, section insights)
  2. Deterministic renderer builds formatted Excel from the plan
  3. Works for ANY combination of tools — not hardcoded to specific artifacts

Usage:
  from ml.api.mbr_report_builder import build_mbr_report

  excel_bytes = await build_mbr_report(
      agent_result=run_mbr_agent_output,
      llm_config={"api_key": "...", "model": "deepseek-chat"},
  )
"""

import json
import io
import re
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ================================================================
# Part 1: STYLE SYSTEM
# ================================================================

class Styles:
    """Centralized style definitions for consistent formatting."""

    # Colors
    NAVY = "1B2A4A"
    WHITE = "FFFFFF"
    LIGHT_GRAY = "F5F6F8"
    MID_GRAY = "E2E4E8"
    DARK_GRAY = "6B7280"
    GREEN = "059669"
    RED = "DC2626"
    AMBER = "D97706"
    BLUE = "2563EB"
    LIGHT_GREEN = "ECFDF5"
    LIGHT_RED = "FEF2F2"
    LIGHT_AMBER = "FFFBEB"
    LIGHT_BLUE = "EFF6FF"
    ACCENT = "4F46E5"

    # Fonts
    TITLE = Font(name="Arial", size=16, bold=True, color=NAVY)
    SUBTITLE = Font(name="Arial", size=11, color=DARK_GRAY)
    SECTION_HEADER = Font(name="Arial", size=13, bold=True, color=NAVY)
    INSIGHT_TEXT = Font(name="Arial", size=10, color=DARK_GRAY, italic=True)
    TABLE_HEADER = Font(name="Arial", size=10, bold=True, color=WHITE)
    TABLE_CELL = Font(name="Arial", size=10, color="111827")
    KPI_VALUE = Font(name="Arial", size=22, bold=True, color=NAVY)
    KPI_LABEL = Font(name="Arial", size=9, color=DARK_GRAY)
    KPI_DELTA = Font(name="Arial", size=9, bold=True)
    NARRATIVE = Font(name="Arial", size=10, color="374151")

    # Fills
    HEADER_FILL = PatternFill("solid", fgColor=NAVY)
    ALT_ROW_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)
    KPI_BG = PatternFill("solid", fgColor=LIGHT_BLUE)
    GREEN_FILL = PatternFill("solid", fgColor=LIGHT_GREEN)
    RED_FILL = PatternFill("solid", fgColor=LIGHT_RED)
    AMBER_FILL = PatternFill("solid", fgColor=LIGHT_AMBER)

    # Borders
    THIN_BORDER = Border(
        bottom=Side(style="thin", color=MID_GRAY),
    )
    TABLE_BORDER = Border(
        bottom=Side(style="thin", color=MID_GRAY),
        left=Side(style="thin", color=MID_GRAY),
        right=Side(style="thin", color=MID_GRAY),
    )

    # Alignment
    WRAP = Alignment(wrap_text=True, vertical="top")
    CENTER = Alignment(horizontal="center", vertical="center")
    RIGHT = Alignment(horizontal="right", vertical="center")


# ================================================================
# Part 2: LLM REPORT PLANNER
# ================================================================

REPORT_PLANNER_PROMPT = """You are a report designer for Monthly Business Reviews.

You receive:
1. A list of available data artifacts (tables) with their labels and column names
2. The executive summary narrative

Your job: design an Excel report by selecting the most important artifacts,
organizing them into logical sheets, choosing appropriate chart types,
and writing a brief insight for each section.

## Output Format — return ONLY valid JSON:

{
  "sheets": [
    {
      "name": "Sheet Name (max 31 chars)",
      "sections": [
        {
          "type": "kpi_cards",
          "metrics": [
            {"label": "Total Revenue", "value_key": "total_revenue", "format": "currency"},
            {"label": "Gross Margin", "value_key": "gross_margin_pct", "format": "percentage"}
          ]
        },
        {
          "type": "narrative",
          "text": "The executive summary text goes here"
        },
        {
          "type": "table",
          "artifact_index": 0,
          "chart": "bar" | "line" | "pie" | "stacked_bar" | null,
          "chart_x_col": "column name for x-axis",
          "chart_y_cols": ["column name(s) for y-axis"],
          "insight": "2-3 sentence interpretation with specific numbers."
        }
      ]
    }
  ]
}

## Rules

1. FIRST SHEET must be "Executive Summary" with kpi_cards + narrative sections.
2. Group related artifacts into logical sheets (Revenue, Margins, Variance, Operations, Anomalies).
3. MAXIMUM 7 sheets total. MAXIMUM 5 tables per sheet. Select only the MOST important artifacts.
4. SKIP: "Column Mapping (verify)", "Detection Config (verify)", duplicate tables, tables with 0-1 rows.
5. Chart selection: time series -> "line", category comparison <= 8 items -> "bar", single metric distribution <= 6 items -> "pie". No charts for 1-2 row or 10+ column tables.
6. Use EXACT column names from the artifact's columns list for chart_x_col and chart_y_cols.
7. Insights must reference SPECIFIC numbers from the data.
8. Format values: "currency" for money, "percentage" for %, "number" for counts, "decimal" for ratios.
9. artifact_index refers to position in the artifacts list (0-indexed).
10. INSIGHT ANTI-HALLUCINATION — CRITICAL:
   - Each artifact below may include a [FACT] line with pre-computed data.
   - Your "insight" field MUST use numbers from [FACT] lines.
   - Do NOT round differently, substitute values, or guess numbers.
   - If no [FACT] is provided, write a generic structural observation — do NOT invent numbers.
   - NEVER claim a category is "largest" or "smallest" unless [FACT] explicitly states the ranking.

RESPOND WITH ONLY VALID JSON. No markdown fences, no explanation.
"""


def generate_insight_from_artifact(artifact):
    """Generate a factual insight string from artifact data.
    Returns a data-grounded sentence. LLM can rephrase but cannot invent numbers."""
    label = artifact.get("label", "")
    data = artifact.get("data", [])
    if not data or not isinstance(data[0], dict):
        return None

    label_lower = label.lower()
    cols = list(data[0].keys())
    n_rows = len(data)

    value_col = None
    group_col = None
    pct_col = None
    for c in cols:
        sample = data[0].get(c)
        cl = c.lower()
        if isinstance(sample, str) and group_col is None:
            group_col = c
        elif isinstance(sample, (int, float)):
            if "pct" in cl or "percent" in cl:
                pct_col = c
            elif value_col is None and cl not in ("count", "transactions", "invoice_count"):
                value_col = c

    if not value_col and not group_col:
        return None

    def _fmt(v):
        if isinstance(v, float):
            if abs(v) >= 1_000_000:
                return f"{v/1_000_000:.1f}M"
            elif abs(v) >= 1_000:
                return f"{v/1_000:.0f}K"
            return f"{v:,.0f}"
        return f"{v:,}" if isinstance(v, int) else str(v)

    try:
        if n_rows == 1:
            row = data[0]
            parts = [f"{k}: {v:,.2f}" if isinstance(v, float) else f"{k}: {v:,}" if isinstance(v, int) else None
                     for k, v in row.items() if k != "metric" and v is not None]
            parts = [p for p in parts if p]
            return f"{label}: {', '.join(parts)}" if parts else None

        if group_col and value_col:
            sorted_data = sorted(data, key=lambda r: r.get(value_col, 0) or 0, reverse=True)
            total = sum(r.get(value_col, 0) or 0 for r in sorted_data if isinstance(r.get(value_col), (int, float)))
            if total == 0:
                return None
            top = sorted_data[0]
            top_name = top.get(group_col, "?")
            top_val = top.get(value_col, 0)
            top_pct = (top_val / total * 100) if total > 0 else 0
            bottom = sorted_data[-1]
            bot_name = bottom.get(group_col, "?")
            bot_val = bottom.get(value_col, 0)
            bot_pct = (bot_val / total * 100) if total > 0 else 0
            insight = f"Highest: {top_name} at {_fmt(top_val)} ({top_pct:.1f}%). Lowest: {bot_name} at {_fmt(bot_val)} ({bot_pct:.1f}%). Total: {_fmt(total)} across {n_rows} items."
            if "margin" in label_lower and "margin_pct" in cols:
                best = max(data, key=lambda r: r.get("margin_pct", 0) or 0)
                worst = min(data, key=lambda r: r.get("margin_pct", 0) or 0)
                insight += f" Best margin: {best.get(group_col, '?')} ({best.get('margin_pct', 0):.1f}%). Worst: {worst.get(group_col, '?')} ({worst.get('margin_pct', 0):.1f}%)."
            return insight

        if "actual" in set(cols) and "target" in set(cols):
            total_actual = sum(r.get("actual", 0) or 0 for r in data)
            total_target = sum(r.get("target", 0) or 0 for r in data)
            if total_target > 0:
                att_pct = total_actual / total_target * 100
                gap = total_actual - total_target
                worst = min(data, key=lambda r: r.get("variance_pct", 0) or 0)
                worst_ctx = " / ".join(str(v) for k, v in worst.items() if k not in ("actual", "target", "variance", "variance_pct") and v is not None)[:60]
                return f"Overall attainment: {att_pct:.1f}% (gap: {gap:+,.0f}). Worst miss: {worst_ctx} at {worst.get('variance_pct', 0):+.1f}%."
    except Exception:
        pass
    return None


def filter_artifacts_for_planner(artifacts, max_artifacts=35):
    """Pre-filter artifacts before sending to LLM report planner.
    Reduces 100+ tables to ~30 by removing noise."""

    SKIP_CONTAINS = ["column mapping", "detection config", "(verify)"]
    SKIP_PREFIXES = ["cross-dim", "composition", "drill-down", "relationship"]

    filtered = []
    seen_labels = set()

    for art in artifacts:
        if art.get("type") != "table":
            continue
        data = art.get("data", [])
        if not data:
            continue

        label = (art.get("label") or "").strip()
        label_lower = label.lower()

        # Skip metadata
        if any(skip in label_lower for skip in SKIP_CONTAINS):
            continue
        # Skip anomaly detail tables
        if any(label_lower.startswith(prefix) for prefix in SKIP_PREFIXES):
            continue
        # Skip outlier detail tables (keep summaries with "count" column)
        if any(kw in label_lower for kw in ("outlier", "z-score", "iqr")):
            if data and isinstance(data[0], dict):
                keys = set(data[0].keys())
                if "row" in keys and ("z_score" in keys or "lower_fence" in keys):
                    continue
                if "count" not in keys:
                    continue
        # Skip waterfall/contribution detail tables
        if "detail" in label_lower and any(kw in label_lower for kw in ("waterfall", "contribution")):
            continue
        # Skip duplicate suffixes: (2), (3), _2, _3
        if re.search(r'\(\d+\)\s*$', label):
            base = re.sub(r'\s*\(\d+\)\s*$', '', label)
            if base in seen_labels:
                continue
        if re.search(r'_\d+$', label):
            base = re.sub(r'_\d+$', '', label)
            if base in seen_labels:
                continue
        # Skip tables with too many rows (raw dumps)
        if len(data) > 50 and not any(kw in label_lower for kw in ("anomaly summary", "negative values", "top findings")):
            continue

        seen_labels.add(label)
        filtered.append(art)
        if len(filtered) >= max_artifacts:
            break

    return filtered


def _build_planner_user_prompt(agent_result):
    """Build the user prompt with artifact list and narrative."""
    all_artifacts = agent_result.get("all_artifacts", [])
    narrative = agent_result.get("narrative", "")
    findings = agent_result.get("findings_chain", [])

    # Pre-filter: reduce to ~30 key tables
    filtered = filter_artifacts_for_planner(all_artifacts)

    lines = ["## Available Artifacts\n"]
    lines.append(f"(Filtered {len(filtered)} key tables from {len(all_artifacts)} total)\n")

    for i, art in enumerate(filtered):
        label = art.get("label", f"Table {i}")
        data = art.get("data", [])
        if not data:
            lines.append(f"[{i}] {label} — EMPTY")
            continue
        cols = list(data[0].keys()) if isinstance(data[0], dict) else []
        n_rows = len(data)
        preview = ""
        if n_rows == 1 and isinstance(data[0], dict):
            preview = " | " + ", ".join(f"{k}={v}" for k, v in data[0].items()
                                         if v is not None)
        lines.append(f"[{i}] {label} — {n_rows} rows, cols: {cols}{preview}")

        # Pre-computed factual insight
        fact = generate_insight_from_artifact(art)
        if fact:
            lines.append(f"    [FACT] {fact}")

    lines.append(f"\n## Executive Summary Narrative\n{narrative[:2000]}")

    lines.append("\n## Key Metrics (from findings)")
    for fc_tool, fc_text in findings:
        if fc_tool in ("kpi_calculation", "margin_analysis", "variance_analysis"):
            lines.append(f"  {fc_tool}: {fc_text[:500]}")

    # Currency detection for planner
    currencies = set()
    for art in filtered:
        for row in (art.get("data") or [])[:5]:
            if isinstance(row, dict):
                cur = row.get("currency")
                if cur and isinstance(cur, str):
                    currencies.add(cur)
    if len(currencies) > 1:
        lines.append(f"\n## CURRENCY WARNING")
        lines.append(f"Data contains multiple currencies: {sorted(currencies)}")
        lines.append(f"Do NOT label totals as any single currency.")

    return "\n".join(lines)


# ================================================================
# Part 3: EXCEL RENDERER
# ================================================================

class MbrExcelRenderer:
    """Renders formatted MBR Excel from a report plan."""

    def __init__(self, artifacts, narrative, result_summary=None):
        self.artifacts = artifacts
        self.narrative = narrative
        self.result_summary = result_summary or {}
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def render(self, plan):
        """Render the full report from plan JSON."""
        sheets = plan.get("sheets", [])
        if not sheets:
            logger.warning("[ReportBuilder] Empty plan, creating fallback")
            self._render_fallback()
            return self._save()

        try:
            for sheet_plan in sheets:
                name = re.sub(r'[:\\/?\*\[\]]', '-', str(sheet_plan.get("name", "Sheet")))[:31]
                ws = self.wb.create_sheet(title=name)
                row = 1

                for section in sheet_plan.get("sections", []):
                    try:
                        section_type = section.get("type")
                        if section_type == "kpi_cards":
                            row = self._render_kpi_cards(ws, row, section)
                        elif section_type == "narrative":
                            row = self._render_narrative(ws, row, section)
                        elif section_type == "table":
                            row = self._render_table_section(ws, row, section)
                        row += 1
                    except Exception as sec_ex:
                        logger.warning(f"[ReportBuilder] Section render failed: {sec_ex}")
                        row += 1

                self._auto_fit_columns(ws)
        except Exception as ex:
            logger.error(f"[ReportBuilder] Plan render failed, using fallback: {ex}")
            # Reset and use fallback
            self.wb = Workbook()
            self.wb.remove(self.wb.active)
            self._render_fallback()

        return self._save()

    def _render_kpi_cards(self, ws, start_row, section):
        """Render KPI metric cards in a horizontal row."""
        metrics = section.get("metrics", [])
        if not metrics:
            return start_row

        row = start_row
        ws.cell(row=row, column=1, value="Key Performance Indicators")
        ws.cell(row=row, column=1).font = Styles.SECTION_HEADER
        row += 1

        for i, metric in enumerate(metrics):
            col = i * 3 + 1
            label = metric.get("label", "")
            value_key = str(metric.get("value_key", ""))
            fmt = metric.get("format", "number")

            value = self.result_summary.get(value_key, "N/A") if value_key else "N/A"
            display_value = self._format_kpi_value(value, fmt)

            for r in range(row, row + 3):
                for c in range(col, col + 2):
                    ws.cell(row=r, column=c).fill = Styles.KPI_BG

            cell = ws.cell(row=row, column=col, value=display_value)
            cell.font = Styles.KPI_VALUE
            cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col + 1)

            cell = ws.cell(row=row + 1, column=col, value=label)
            cell.font = Styles.KPI_LABEL
            cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=row + 1, start_column=col,
                           end_row=row + 1, end_column=col + 1)

        return row + 3

    def _render_narrative(self, ws, start_row, section):
        """Render narrative text with markdown-like formatting."""
        text = section.get("text") or self.narrative or ""
        if not text:
            return start_row

        row = start_row
        ws.cell(row=row, column=1, value="Executive Summary")
        ws.cell(row=row, column=1).font = Styles.SECTION_HEADER
        row += 1

        for para in text.split("\n"):
            para = para.strip()
            if not para:
                continue

            if para.startswith("##"):
                para = para.lstrip("#").strip()
                cell = ws.cell(row=row, column=1, value=para)
                cell.font = Font(name="Arial", size=11, bold=True, color=Styles.NAVY)
            elif para.startswith("**[P"):
                clean = re.sub(r'\*\*', '', para)
                cell = ws.cell(row=row, column=1, value=clean)
                cell.font = Font(name="Arial", size=10, bold=True, color=Styles.ACCENT)
            elif para.startswith("*") or para.startswith("-"):
                clean = para.lstrip("*- ").strip()
                cell = ws.cell(row=row, column=1, value=f"  \u2022  {clean}")
                cell.font = Styles.NARRATIVE
            else:
                clean = re.sub(r'\*\*([^*]+)\*\*', r'\1', para)
                cell = ws.cell(row=row, column=1, value=clean)
                cell.font = Styles.NARRATIVE

            cell.alignment = Styles.WRAP
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            row += 1

        return row

    def _render_table_section(self, ws, start_row, section):
        """Render a data table with optional chart and insight."""
        idx = section.get("artifact_index")
        if idx is None:
            return start_row
        try:
            idx = int(idx)
        except (ValueError, TypeError):
            return start_row
        if idx < 0 or idx >= len(self.artifacts):
            return start_row

        artifact = self.artifacts[idx]
        data = artifact.get("data", [])
        if not data or not isinstance(data[0], dict):
            return start_row

        label = artifact.get("label", f"Table {idx}")
        insight = section.get("insight")
        chart_type = section.get("chart")
        chart_x = section.get("chart_x_col")
        chart_y = section.get("chart_y_cols", [])

        row = start_row

        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Styles.SECTION_HEADER
        row += 1

        if insight:
            cell = ws.cell(row=row, column=1, value=insight)
            cell.font = Styles.INSIGHT_TEXT
            cell.alignment = Styles.WRAP
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            row += 1

        row += 1

        columns = list(data[0].keys())
        for c_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row, column=c_idx, value=col_name)
            cell.font = Styles.TABLE_HEADER
            cell.fill = Styles.HEADER_FILL
            cell.alignment = Styles.CENTER
            cell.border = Styles.TABLE_BORDER

        header_row = row
        row += 1

        data_start_row = row
        for r_idx, record in enumerate(data[:50]):
            for c_idx, col_name in enumerate(columns, 1):
                value = record.get(col_name)
                cell = ws.cell(row=row, column=c_idx)

                if value is None:
                    cell.value = ""
                elif isinstance(value, float) and (value != value):
                    cell.value = ""  # NaN
                elif isinstance(value, (int, float)):
                    cell.value = value
                    col_lower = col_name.lower()
                    if "pct" in col_lower or "percent" in col_lower:
                        cell.number_format = '0.0"%"'
                    elif any(kw in col_lower for kw in ("revenue", "amount", "cost", "margin",
                                                         "profit", "value", "target", "cogs")):
                        cell.number_format = '#,##0.00'
                    elif isinstance(value, float):
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '#,##0'
                else:
                    cell.value = str(value)

                cell.font = Styles.TABLE_CELL
                cell.border = Styles.THIN_BORDER

                if isinstance(value, (int, float)):
                    if "variance" in col_name.lower() or "delta" in col_name.lower():
                        if value < 0:
                            cell.font = Font(name="Arial", size=10, color=Styles.RED)
                            cell.fill = Styles.RED_FILL
                        elif value > 0:
                            cell.font = Font(name="Arial", size=10, color=Styles.GREEN)
                            cell.fill = Styles.GREEN_FILL

                cell.alignment = Styles.RIGHT if isinstance(value, (int, float)) else \
                    Alignment(vertical="center")

            if r_idx % 2 == 1:
                for c_idx in range(1, len(columns) + 1):
                    c = ws.cell(row=row, column=c_idx)
                    if not c.fill.fgColor or c.fill.fgColor.rgb in ("00000000", Styles.WHITE):
                        c.fill = Styles.ALT_ROW_FILL

            row += 1

        data_end_row = row - 1

        if len(data) > 50:
            cell = ws.cell(row=row, column=1,
                           value=f"... showing 50 of {len(data)} rows")
            cell.font = Font(name="Arial", size=9, color=Styles.DARK_GRAY, italic=True)
            row += 1

        if chart_type and len(data) >= 2 and chart_x and chart_y:
            row = self._render_chart(ws, row, chart_type, chart_x, chart_y,
                                     columns, header_row, data_start_row,
                                     data_end_row, label)
            row += 1

        return row + 1

    def _render_chart(self, ws, start_row, chart_type, x_col, y_cols,
                      columns, header_row, data_start_row, data_end_row, title):
        """Add a chart below the table."""
        try:
            x_idx = columns.index(x_col) + 1 if x_col in columns else None
            y_indices = [columns.index(yc) + 1 for yc in y_cols if yc in columns]

            if not x_idx or not y_indices:
                return start_row

            if chart_type == "pie":
                chart = PieChart()
                chart.style = 10
            elif chart_type == "line":
                chart = LineChart()
                chart.style = 10
                chart.y_axis.numFmt = '#,##0'
            elif chart_type == "stacked_bar":
                chart = BarChart()
                chart.type = "col"
                chart.grouping = "stacked"
                chart.style = 10
                chart.y_axis.numFmt = '#,##0'
            else:
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.y_axis.numFmt = '#,##0'

            chart.title = title
            chart.width = 18
            chart.height = 11

            cats = Reference(ws, min_col=x_idx, min_row=data_start_row,
                             max_row=data_end_row)

            for yi in y_indices:
                values = Reference(ws, min_col=yi, min_row=header_row,
                                   max_row=data_end_row)
                chart.add_data(values, titles_from_data=True)

            chart.set_categories(cats)

            colors = ["4F46E5", "10B981", "F59E0B", "EF4444", "8B5CF6", "06B6D4"]
            for i, series in enumerate(chart.series):
                series.graphicalProperties.solidFill = colors[i % len(colors)]

            ws.add_chart(chart, f"A{start_row + 1}")
            return start_row + 16

        except Exception as e:
            logger.warning(f"[ReportBuilder] Chart failed: {e}")
            return start_row

    def _format_kpi_value(self, value, fmt):
        """Format a KPI value for display."""
        if value is None or value == "N/A":
            return "N/A"
        try:
            v = float(value)
            if fmt == "currency":
                if abs(v) >= 1_000_000:
                    return f"${v / 1_000_000:.1f}M"
                elif abs(v) >= 1_000:
                    return f"${v / 1_000:.0f}K"
                return f"${v:,.0f}"
            elif fmt == "percentage":
                return f"{v:.1f}%"
            elif fmt == "decimal":
                return f"{v:.2f}"
            else:
                if abs(v) >= 1_000_000:
                    return f"{v / 1_000_000:.1f}M"
                return f"{v:,.0f}"
        except (ValueError, TypeError):
            return str(value)

    def _auto_fit_columns(self, ws):
        """Auto-fit column widths based on content."""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    cell_len = len(str(cell.value))
                    max_len = max(max_len, min(cell_len, 40))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    def _render_fallback(self):
        """Fallback: dump key artifacts with basic formatting."""
        ws = self.wb.create_sheet(title="MBR Report")
        row = 1
        ws.cell(row=row, column=1, value="Monthly Business Review")
        ws.cell(row=row, column=1).font = Styles.TITLE
        row += 2

        if self.narrative:
            for line in self.narrative.split("\n")[:50]:
                line = line.strip()
                if line:
                    ws.cell(row=row, column=1, value=re.sub(r'\*\*', '', line))
                    ws.cell(row=row, column=1).font = Styles.NARRATIVE
                    row += 1
            row += 2

        for art in self.artifacts[:10]:
            if art.get("type") != "table" or not art.get("data"):
                continue
            label = art.get("label", "Table")
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Styles.SECTION_HEADER
            row += 1
            data = art["data"][:20]
            if data and isinstance(data[0], dict):
                cols = list(data[0].keys())
                for ci, cn in enumerate(cols, 1):
                    ws.cell(row=row, column=ci, value=cn)
                    ws.cell(row=row, column=ci).font = Styles.TABLE_HEADER
                    ws.cell(row=row, column=ci).fill = Styles.HEADER_FILL
                row += 1
                for record in data:
                    for ci, cn in enumerate(cols, 1):
                        v = record.get(cn)
                        if isinstance(v, float) and v != v:
                            v = None
                        ws.cell(row=row, column=ci, value=v)
                        ws.cell(row=row, column=ci).font = Styles.TABLE_CELL
                    row += 1
            row += 2

        self._auto_fit_columns(ws)

    def _save(self):
        """Save workbook to bytes."""
        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf.getvalue()


# ================================================================
# Part 4: MAIN ENTRY POINT
# ================================================================

async def build_mbr_report(agent_result, llm_config, call_llm_fn=None):
    """
    Build formatted MBR Excel report.

    Args:
        agent_result: output from run_mbr_agent()
        llm_config: {"api_key", "model", "base_url"}
        call_llm_fn: async fn(prompt, system_prompt, config) -> str

    Returns:
        bytes — Excel file content
    """
    artifacts = agent_result.get("all_artifacts", [])
    narrative = agent_result.get("narrative", "")

    # Extract result_summary from findings_chain text + artifacts
    result_summary = {}
    for tool_id, text in agent_result.get("findings_chain", []):
        if tool_id == "kpi_calculation":
            for line in text.split("\n"):
                # Parse pipe-delimited: "total_revenue: 12,155,895.09 | margin_pct: 56.3"
                for part in line.split("|"):
                    part = part.strip()
                    if ":" in part and not part.startswith("Top") and not part.startswith("  "):
                        key, val = part.split(":", 1)
                        key = key.strip()
                        val = val.strip().replace(",", "")
                        try:
                            result_summary[key] = float(val)
                        except ValueError:
                            result_summary[key] = val
                # Parse "Target attainment: 61.7% (actual=9,341,903 vs target=15,144,479, gap=-5,802,576)"
                if "target attainment:" in line.lower():
                    import re as _re
                    m = _re.search(r'(\d+\.?\d*)%', line)
                    if m:
                        result_summary["target_attainment_pct"] = float(m.group(1))
                    m = _re.search(r'gap=([+-]?[\d,]+)', line)
                    if m:
                        result_summary["target_gap"] = float(m.group(1).replace(",", ""))

    # Also extract from target variance artifacts directly
    for art in agent_result.get("all_artifacts", []):
        data = art.get("data", [])
        if not data or not isinstance(data[0], dict):
            continue
        keys = set(data[0].keys())
        if "actual" in keys and "target" in keys and "target_attainment_pct" not in result_summary:
            total_actual = sum(r.get("actual", 0) or 0 for r in data)
            total_target = sum(r.get("target", 0) or 0 for r in data)
            if total_target > 0:
                result_summary["target_attainment_pct"] = round(total_actual / total_target * 100, 1)
                result_summary["target_gap"] = total_actual - total_target

    # Step 1: LLM plans the report
    user_prompt = _build_planner_user_prompt(agent_result)

    # Append available metric keys so planner doesn't guess non-existent ones
    if result_summary:
        user_prompt += "\n\n## Available KPI Metric Keys (for kpi_cards value_key)\n"
        user_prompt += "ONLY use these keys. Do NOT invent keys that are not listed.\n"
        for k, v in result_summary.items():
            user_prompt += f"  {k}: {v}\n"

    plan = None
    try:
        if call_llm_fn:
            raw = await call_llm_fn(user_prompt, REPORT_PLANNER_PROMPT, llm_config)
        else:
            import httpx
            import os
            api_key = llm_config.get("api_key") or os.getenv("DEEPSEEK_API_KEY")
            base_url = llm_config.get("base_url") or os.getenv("DEEPSEEK_BASE_URL",
                                                                 "https://api.deepseek.com")
            model = llm_config.get("model", "deepseek-chat")

            url = f"{base_url}/chat/completions"
            headers = {"Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json"}
            payload = {
                "model": model, "temperature": 0.1, "max_tokens": 4000,
                "messages": [
                    {"role": "system", "content": REPORT_PLANNER_PROMPT},
                    {"role": "user", "content": user_prompt},
                ],
            }
            async with httpx.AsyncClient(timeout=120) as client:
                resp = await client.post(url, json=payload, headers=headers)
                resp.raise_for_status()
                data = resp.json()
                raw = data["choices"][0]["message"].get("content", "")

        raw = raw.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        s = raw.find("{")
        e = raw.rfind("}")
        if s != -1 and e != -1:
            plan = json.loads(raw[s:e + 1])
            logger.info(f"[ReportBuilder] LLM plan: {len(plan.get('sheets', []))} sheets")
        else:
            logger.warning(f"[ReportBuilder] No JSON found in LLM response: {raw[:200]}")
            plan = None

    except json.JSONDecodeError as jex:
        logger.error(f"[ReportBuilder] JSON parse failed: {jex}. Raw: {raw[:200] if 'raw' in dir() else '?'}")
        plan = None
    except Exception as ex:
        logger.error(f"[ReportBuilder] LLM planning failed: {type(ex).__name__}: {ex}")
        plan = None

    # Step 2: Render Excel (use filtered artifacts so artifact_index matches planner output)
    filtered = filter_artifacts_for_planner(artifacts)
    renderer = MbrExcelRenderer(filtered, narrative, result_summary)
    excel_bytes = renderer.render(plan or {})

    return excel_bytes
