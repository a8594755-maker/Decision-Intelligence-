#!/usr/bin/env python3
"""
Generate 12 months of realistic company reports for testing the Style Learning pipeline.
Period: 2025-03 to 2026-03

Output types:
  - Monthly MBR (Excel, 6-8 sheets, KPIs, charts placeholder, dashboard)
  - Weekly Ops Summary (Excel, 2-3 sheets, concise)
  - Quarterly QBR Deck (Excel, 8-10 sheets, executive level)
  - Risk Report (Excel, 3-4 sheets, exception log)
  - Forecast Report (Excel, 3 sheets, demand/supply)

All files use consistent company style:
  - Header: dark indigo bg (#1F3864), white text, 微軟正黑體/Calibri
  - KPI naming: Chinese name + English abbreviation
  - Alternating row colors
  - Freeze panes on data sheets
  - Realistic supply chain KPIs with monthly variance
"""

import os
import random
import math
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'seed-data')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Company Style Constants ──────────────────────────────────
HEADER_FILL = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
SUBHEADER_FILL = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
SUBHEADER_FONT = Font(name='Calibri', bold=True, color='1F3864', size=10)
ALT_ROW_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
KPI_GOOD_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
KPI_BAD_FILL = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
KPI_NEUTRAL_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
TITLE_FONT = Font(name='Calibri', bold=True, color='1F3864', size=16)
BODY_FONT = Font(name='Calibri', size=10)
NUM_FONT = Font(name='Calibri', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

# ── Data Generation Helpers ──────────────────────────────────
PRODUCTS = ['Widget-A', 'Widget-B', 'Sensor-X', 'Module-M1', 'Board-PCB3', 'Cap-220uF', 'IC-7805', 'Relay-K1']
PLANTS = ['TW-FAB1', 'TW-FAB2', 'SG-PLANT', 'VN-ASSEMBLY']
SUPPLIERS = ['Foxconn', 'Delta Electronics', 'Murata', 'TDK', 'Samsung SDI', 'TSMC', 'ASE Group', 'Yageo']
REGIONS = ['台灣北區', '台灣中區', '台灣南區', '新加坡', '越南']
RISK_TYPES = ['供應中斷', '品質異常', '交期延遲', '價格波動', '產能不足', '庫存過高', '匯率風險']
ISSUE_CATEGORIES = ['Material Shortage', 'Quality Defect', 'Logistics Delay', 'Capacity Constraint', 'Demand Spike', 'System Error']

def seasonal_factor(month):
    """Simulate seasonal demand pattern"""
    return 1.0 + 0.15 * math.sin((month - 3) * math.pi / 6)

def gen_kpi_value(base, variance=0.1, month=1):
    """Generate a KPI value with seasonal variance"""
    seasonal = seasonal_factor(month)
    noise = random.gauss(0, variance * base)
    return max(0, base * seasonal + noise)

def gen_pct(base, variance=0.05):
    return min(1.0, max(0.0, base + random.gauss(0, variance)))

def format_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def format_data_rows(ws, start_row, end_row, col_count):
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if (r - start_row) % 2 == 1:
                cell.fill = ALT_ROW_FILL

def auto_width(ws, col_count, min_width=10):
    for c in range(1, col_count + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(min_width, 14)

def add_cover_sheet(wb, title, period, author='Supply Chain Analytics Team'):
    ws = wb.active
    ws.title = '封面 Cover'
    ws.merge_cells('B3:F3')
    ws['B3'] = title
    ws['B3'].font = Font(name='Calibri', bold=True, color='1F3864', size=22)
    ws['B5'] = f'報告期間：{period}'
    ws['B5'].font = Font(name='Calibri', size=12, color='333333')
    ws['B7'] = f'製作單位：{author}'
    ws['B7'].font = Font(name='Calibri', size=11, color='666666')
    ws['B8'] = f'產出日期：{datetime.now().strftime("%Y-%m-%d")}'
    ws['B8'].font = Font(name='Calibri', size=11, color='666666')
    ws['B10'] = '本報告由 Digital Worker 自動產出，內容經主管審核。'
    ws['B10'].font = Font(name='Calibri', size=10, italic=True, color='999999')
    ws.sheet_properties.tabColor = '1F3864'

# ── MBR Report Generator ────────────────────────────────────
def generate_mbr(year, month):
    wb = Workbook()
    period = f'{year}年{month:02d}月'
    dt = datetime(year, month, 1)

    # 1. Cover
    add_cover_sheet(wb, f'月營運報告 Monthly Business Review', period)

    # 2. KPI Dashboard
    ws = wb.create_sheet('KPI Dashboard')
    ws.sheet_properties.tabColor = '2E75B6'
    kpis = [
        ('需求預測準確率 MAPE', f'{gen_pct(0.88, 0.04):.1%}', '≥ 85%'),
        ('訂單達交率 OTD', f'{gen_pct(0.92, 0.03):.1%}', '≥ 90%'),
        ('庫存周轉天數 ITO', f'{gen_kpi_value(45, 5, month):.0f} days', '≤ 50 days'),
        ('服務水準 Service Level', f'{gen_pct(0.95, 0.02):.1%}', '≥ 93%'),
        ('缺料率 Shortage Rate', f'{gen_pct(0.03, 0.015):.1%}', '≤ 5%'),
        ('成本節省 Cost Savings', f'${gen_kpi_value(120000, 30000, month):,.0f}', '≥ $100K'),
        ('供應商準時率 Supplier OTD', f'{gen_pct(0.89, 0.05):.1%}', '≥ 85%'),
        ('品質退貨率 Return Rate', f'{gen_pct(0.012, 0.005):.2%}', '≤ 2%'),
    ]
    ws['B2'] = f'關鍵績效指標 Key Performance Indicators — {period}'
    ws['B2'].font = Font(name='Calibri', bold=True, color='1F3864', size=14)
    headers = ['指標名稱 KPI', '本月實績 Actual', '目標 Target', '狀態 Status']
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c+1, value=h)
    format_header_row(ws, 4, len(headers) + 1)
    for i, (name, actual, target) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=actual)
        ws.cell(row=r, column=4, value=target)
        # Simple status
        status = random.choice(['達標', '達標', '達標', '未達標', '觀察中'])
        ws.cell(row=r, column=5, value=status)
        fill = KPI_GOOD_FILL if status == '達標' else KPI_BAD_FILL if status == '未達標' else KPI_NEUTRAL_FILL
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = BODY_FONT
    auto_width(ws, 5, 18)

    # 3. Cleaned Data
    ws = wb.create_sheet('Cleaned_Data')
    ws.sheet_properties.tabColor = '548235'
    data_headers = ['Date', 'Product', 'Plant', 'Demand_Qty', 'Supply_Qty', 'Inventory', 'Backlog', 'Unit_Cost']
    for c, h in enumerate(data_headers, 1):
        ws.cell(row=1, column=c, value=h)
    format_header_row(ws, 1, len(data_headers))
    ws.freeze_panes = 'A2'
    row = 2
    for day in range(1, 29):
        for prod in random.sample(PRODUCTS, 4):
            plant = random.choice(PLANTS)
            demand = int(gen_kpi_value(500, 100, month))
            supply = int(demand * gen_pct(0.95, 0.08))
            inventory = int(gen_kpi_value(2000, 400, month))
            backlog = max(0, demand - supply)
            cost = round(gen_kpi_value(12.5, 2, month), 2)
            ws.cell(row=row, column=1, value=datetime(year, month, min(day, 28)).strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=prod)
            ws.cell(row=row, column=3, value=plant)
            ws.cell(row=row, column=4, value=demand)
            ws.cell(row=row, column=5, value=supply)
            ws.cell(row=row, column=6, value=inventory)
            ws.cell(row=row, column=7, value=backlog)
            ws.cell(row=row, column=8, value=cost)
            row += 1
    format_data_rows(ws, 2, row - 1, len(data_headers))
    auto_width(ws, len(data_headers))

    # 4. Data Issues
    ws = wb.create_sheet('Data_Issues')
    ws.sheet_properties.tabColor = 'C00000'
    issue_headers = ['Issue ID', 'Category', 'Description', 'Severity', 'Status', 'Owner', 'Detected Date']
    for c, h in enumerate(issue_headers, 1):
        ws.cell(row=1, column=c, value=h)
    format_header_row(ws, 1, len(issue_headers))
    num_issues = random.randint(3, 8)
    for i in range(num_issues):
        r = 2 + i
        ws.cell(row=r, column=1, value=f'DI-{year}{month:02d}-{i+1:03d}')
        ws.cell(row=r, column=2, value=random.choice(ISSUE_CATEGORIES))
        ws.cell(row=r, column=3, value=random.choice([
            '原物料交期延遲超過 5 天', '系統資料缺失，需人工補錄',
            '供應商品質數據異常，已通知 SQE 確認', '庫存數量與 ERP 不一致',
            '需求預測偏差超過閾值', '運輸途中貨物受損',
        ]))
        ws.cell(row=r, column=4, value=random.choice(['High', 'Medium', 'Low']))
        ws.cell(row=r, column=5, value=random.choice(['Open', 'In Progress', 'Resolved', 'Closed']))
        ws.cell(row=r, column=6, value=random.choice(['王大明', '李小華', '陳志偉', 'Alex Chen']))
        ws.cell(row=r, column=7, value=dt.strftime('%Y-%m-%d'))
    format_data_rows(ws, 2, 2 + num_issues - 1, len(issue_headers))
    auto_width(ws, len(issue_headers), 14)

    # 5. Forecast vs Actual
    ws = wb.create_sheet('Forecast')
    ws.sheet_properties.tabColor = 'ED7D31'
    fc_headers = ['Product', 'Forecast_Qty', 'Actual_Qty', 'Variance', 'MAPE', 'Bias']
    for c, h in enumerate(fc_headers, 1):
        ws.cell(row=1, column=c, value=h)
    format_header_row(ws, 1, len(fc_headers))
    ws.freeze_panes = 'A2'
    for i, prod in enumerate(PRODUCTS):
        r = 2 + i
        fc = int(gen_kpi_value(15000, 3000, month))
        actual = int(fc * gen_pct(1.0, 0.12))
        variance = actual - fc
        mape = abs(variance) / max(actual, 1)
        bias = variance / max(actual, 1)
        ws.cell(row=r, column=1, value=prod)
        ws.cell(row=r, column=2, value=fc)
        ws.cell(row=r, column=3, value=actual)
        ws.cell(row=r, column=4, value=variance)
        ws.cell(row=r, column=5, value=round(mape, 3))
        ws.cell(row=r, column=5).number_format = '0.0%'
        ws.cell(row=r, column=6, value=round(bias, 3))
        ws.cell(row=r, column=6).number_format = '0.0%'
    format_data_rows(ws, 2, 2 + len(PRODUCTS) - 1, len(fc_headers))
    auto_width(ws, len(fc_headers))

    # 6. Plan Summary
    ws = wb.create_sheet('Plan')
    ws.sheet_properties.tabColor = '4472C4'
    plan_headers = ['Product', 'Plant', 'Planned_Qty', 'Safety_Stock', 'Reorder_Point', 'Lead_Time_Days', 'Supplier']
    for c, h in enumerate(plan_headers, 1):
        ws.cell(row=1, column=c, value=h)
    format_header_row(ws, 1, len(plan_headers))
    row = 2
    for prod in PRODUCTS:
        for plant in random.sample(PLANTS, 2):
            ws.cell(row=row, column=1, value=prod)
            ws.cell(row=row, column=2, value=plant)
            ws.cell(row=row, column=3, value=int(gen_kpi_value(5000, 1000, month)))
            ws.cell(row=row, column=4, value=int(gen_kpi_value(800, 200, month)))
            ws.cell(row=row, column=5, value=int(gen_kpi_value(1200, 300, month)))
            ws.cell(row=row, column=6, value=random.choice([7, 14, 21, 30, 45]))
            ws.cell(row=row, column=7, value=random.choice(SUPPLIERS))
            row += 1
    format_data_rows(ws, 2, row - 1, len(plan_headers))
    auto_width(ws, len(plan_headers))

    # 7. Analysis Summary
    ws = wb.create_sheet('Analysis')
    ws.sheet_properties.tabColor = '7030A0'
    ws['B2'] = f'月度分析摘要 Monthly Analysis Summary — {period}'
    ws['B2'].font = Font(name='Calibri', bold=True, color='1F3864', size=13)
    observations = [
        f'- 本月整體需求較上月{"增加" if random.random() > 0.5 else "減少"} {random.randint(3, 15)}%，主要受 {random.choice(PRODUCTS)} 拉動。',
        f'- {random.choice(SUPPLIERS)} 交期延遲問題持續，已啟動備援供應商評估。',
        f'- 庫存周轉天數為 {random.randint(38, 55)} 天，{"符合" if random.random() > 0.4 else "略高於"} 目標。',
        f'- {random.choice(PLANTS)} 產能利用率達 {random.randint(78, 96)}%，建議關注瓶頸工站。',
        f'- 預測準確率 MAPE 為 {random.randint(8, 18)}%，{random.choice(PRODUCTS)} 偏差較大，已調整模型參數。',
        f'- 本月共處理 {random.randint(5, 20)} 件資料品質問題，{random.randint(70, 95)}% 已解決。',
        f'- 建議下月重點：{"強化安全庫存策略" if random.random() > 0.5 else "優化供應商交期管理"}。',
    ]
    for i, obs in enumerate(observations):
        ws[f'B{4+i}'] = obs
        ws[f'B{4+i}'].font = Font(name='Calibri', size=10, color='333333')
    auto_width(ws, 2, 80)

    fname = f'MBR_{year}{month:02d}_月營運報告.xlsx'
    wb.save(os.path.join(OUTPUT_DIR, fname))
    return fname

# ── Weekly Ops Summary Generator ─────────────────────────────
def generate_weekly_ops(year, month, week):
    wb = Workbook()
    dt = datetime(year, month, min(week * 7, 28))
    period = f'{year}/{month:02d} W{week}'

    ws = wb.active
    ws.title = '週報摘要 Summary'
    ws['B2'] = f'週營運摘要 Weekly Ops Summary — {period}'
    ws['B2'].font = TITLE_FONT

    summary_items = [
        f'本週出貨量：{int(gen_kpi_value(8000, 1500, month)):,} units',
        f'訂單達交率：{gen_pct(0.93, 0.03):.1%}',
        f'新增訂單：{random.randint(40, 120)} 筆',
        f'異常事件：{random.randint(0, 5)} 件',
        f'供應商交期達成：{gen_pct(0.88, 0.05):.1%}',
    ]
    for i, item in enumerate(summary_items):
        ws[f'B{4+i}'] = item
        ws[f'B{4+i}'].font = BODY_FONT

    # Action items
    ws[f'B{10}'] = '待辦事項 Action Items'
    ws[f'B{10}'].font = SUBHEADER_FONT
    actions = [
        f'1. 跟進 {random.choice(SUPPLIERS)} 延遲訂單（負責人：{random.choice(["王大明", "李小華"])}）',
        f'2. 確認 {random.choice(PRODUCTS)} 安全庫存水位',
        f'3. 更新下週需求預測',
    ]
    for i, act in enumerate(actions):
        ws[f'B{11+i}'] = act
        ws[f'B{11+i}'].font = BODY_FONT

    # Data sheet
    ws2 = wb.create_sheet('Detail')
    detail_headers = ['Day', 'Shipments', 'Orders', 'Backlog', 'Issues']
    for c, h in enumerate(detail_headers, 1):
        ws2.cell(row=1, column=c, value=h)
    format_header_row(ws2, 1, len(detail_headers))
    for d in range(5):
        r = 2 + d
        ws2.cell(row=r, column=1, value=f'Day {d+1}')
        ws2.cell(row=r, column=2, value=int(gen_kpi_value(1600, 300, month)))
        ws2.cell(row=r, column=3, value=random.randint(8, 25))
        ws2.cell(row=r, column=4, value=random.randint(0, 200))
        ws2.cell(row=r, column=5, value=random.randint(0, 3))
    format_data_rows(ws2, 2, 6, len(detail_headers))
    auto_width(ws2, len(detail_headers))

    fname = f'週報_{year}{month:02d}_W{week}_Weekly_Ops.xlsx'
    wb.save(os.path.join(OUTPUT_DIR, fname))
    return fname

# ── QBR Report Generator ────────────────────────────────────
def generate_qbr(year, quarter):
    wb = Workbook()
    q_months = {1: (1,3), 2: (4,6), 3: (7,9), 4: (10,12)}
    start_m, end_m = q_months[quarter]
    period = f'{year} Q{quarter} ({start_m}月-{end_m}月)'

    add_cover_sheet(wb, f'季度營運報告 Quarterly Business Review', period)

    # Executive Summary
    ws = wb.create_sheet('Executive Summary')
    ws.sheet_properties.tabColor = '1F3864'
    ws['B2'] = f'執行摘要 Executive Summary — Q{quarter} {year}'
    ws['B2'].font = Font(name='Calibri', bold=True, color='1F3864', size=16)
    exec_points = [
        f'- 本季營收達成率 {gen_pct(0.96, 0.04):.1%}，{"超越" if random.random() > 0.4 else "略低於"}年度目標。',
        f'- 供應鏈總成本較去年同期{"下降" if random.random() > 0.5 else "上升"} {random.randint(2, 8)}%。',
        f'- 平均訂單達交率 {gen_pct(0.91, 0.03):.1%}，客戶滿意度維持高檔。',
        f'- 庫存周轉率 {gen_kpi_value(8.5, 1.5, quarter*3):.1f}x，持續優化中。',
        f'- 重大風險事件 {random.randint(1, 4)} 件，均已妥善處置。',
        f'- 下季重點：{"推動供應商整合" if random.random() > 0.5 else "加速數位轉型專案"}。',
    ]
    for i, pt in enumerate(exec_points):
        ws[f'B{4+i}'] = pt
        ws[f'B{4+i}'].font = Font(name='Calibri', size=11, color='333333')

    # Quarterly KPIs
    ws = wb.create_sheet('Quarterly KPIs')
    ws.sheet_properties.tabColor = '2E75B6'
    q_kpi_headers = ['指標 KPI', 'Q{} Actual'.format(quarter), 'Target', 'YoY Change', 'Status']
    for c, h in enumerate(q_kpi_headers, 1):
        ws.cell(row=1, column=c, value=h)
    format_header_row(ws, 1, len(q_kpi_headers))
    q_kpis = [
        ('營收達成率 Revenue Achievement', f'{gen_pct(0.96, 0.04):.1%}', '95%', f'{random.choice(["+", "-"])}{random.randint(1,6)}%'),
        ('供應鏈成本 SC Cost', f'${gen_kpi_value(2800000, 500000, quarter*3):,.0f}', '$3.0M', f'{random.choice(["+", "-"])}{random.randint(1,8)}%'),
        ('平均 MAPE', f'{gen_pct(0.12, 0.03):.1%}', '≤ 15%', f'{random.choice(["+", "-"])}{random.randint(0,3)}pp'),
        ('客訴件數 Complaints', f'{random.randint(5, 25)}', '≤ 20', f'{random.choice(["+", "-"])}{random.randint(1,10)}'),
        ('ESG 碳排放 (tCO2)', f'{gen_kpi_value(450, 80, quarter*3):,.0f}', '≤ 500', f'{random.choice(["+", "-"])}{random.randint(2,12)}%'),
    ]
    for i, (name, actual, target, yoy) in enumerate(q_kpis):
        r = 2 + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=actual)
        ws.cell(row=r, column=3, value=target)
        ws.cell(row=r, column=4, value=yoy)
        ws.cell(row=r, column=5, value=random.choice(['達標', '達標', '未達標', '觀察中']))
    format_data_rows(ws, 2, 2 + len(q_kpis) - 1, len(q_kpi_headers))
    auto_width(ws, len(q_kpi_headers), 16)

    # Trend (monthly breakdown)
    ws = wb.create_sheet('Monthly Trend')
    trend_headers = ['Month', 'Revenue ($K)', 'Orders', 'OTD %', 'MAPE %', 'Inventory Days']
    for c, h in enumerate(trend_headers, 1):
        ws.cell(row=1, column=c, value=h)
    format_header_row(ws, 1, len(trend_headers))
    for i, m in enumerate(range(start_m, end_m + 1)):
        r = 2 + i
        ws.cell(row=r, column=1, value=f'{year}/{m:02d}')
        ws.cell(row=r, column=2, value=int(gen_kpi_value(950, 150, m)))
        ws.cell(row=r, column=3, value=int(gen_kpi_value(350, 60, m)))
        ws.cell(row=r, column=4, value=round(gen_pct(0.92, 0.03), 3))
        ws.cell(row=r, column=5, value=round(gen_pct(0.12, 0.03), 3))
        ws.cell(row=r, column=6, value=int(gen_kpi_value(45, 5, m)))
    format_data_rows(ws, 2, 2 + 2, len(trend_headers))
    auto_width(ws, len(trend_headers))

    # Risk Events
    ws = wb.create_sheet('Risk Events')
    ws.sheet_properties.tabColor = 'C00000'
    risk_headers = ['Event ID', 'Date', 'Risk Type', 'Description', 'Impact Level', 'Response', 'Status']
    for c, h in enumerate(risk_headers, 1):
        ws.cell(row=1, column=c, value=h)
    format_header_row(ws, 1, len(risk_headers))
    for i in range(random.randint(2, 6)):
        r = 2 + i
        ws.cell(row=r, column=1, value=f'RE-{year}Q{quarter}-{i+1:02d}')
        ws.cell(row=r, column=2, value=f'{year}/{random.choice(range(start_m, end_m+1)):02d}/{random.randint(1,28):02d}')
        ws.cell(row=r, column=3, value=random.choice(RISK_TYPES))
        ws.cell(row=r, column=4, value=random.choice([
            f'{random.choice(SUPPLIERS)} 原料供應延遲 {random.randint(3,14)} 天',
            f'{random.choice(PLANTS)} 設備故障停機 {random.randint(4,24)} 小時',
            f'{random.choice(PRODUCTS)} 客訴品質異常，批次追溯中',
        ]))
        ws.cell(row=r, column=5, value=random.choice(['Critical', 'High', 'Medium']))
        ws.cell(row=r, column=6, value=random.choice(['已啟動備援', '持續監控', '已解決', '升級處理']))
        ws.cell(row=r, column=7, value=random.choice(['Open', 'Closed', 'Monitoring']))
    format_data_rows(ws, 2, r, len(risk_headers))
    auto_width(ws, len(risk_headers), 14)

    # Supplier Scorecard
    ws = wb.create_sheet('Supplier Scorecard')
    ws.sheet_properties.tabColor = '548235'
    sc_headers = ['Supplier', 'OTD %', 'Quality PPM', 'Cost Index', 'Lead Time', 'Overall Score', 'Trend']
    for c, h in enumerate(sc_headers, 1):
        ws.cell(row=1, column=c, value=h)
    format_header_row(ws, 1, len(sc_headers))
    for i, sup in enumerate(SUPPLIERS):
        r = 2 + i
        ws.cell(row=r, column=1, value=sup)
        ws.cell(row=r, column=2, value=round(gen_pct(0.90, 0.06), 3))
        ws.cell(row=r, column=3, value=random.randint(50, 800))
        ws.cell(row=r, column=4, value=round(gen_pct(1.0, 0.08), 2))
        ws.cell(row=r, column=5, value=f'{random.choice([7,14,21,30])} days')
        ws.cell(row=r, column=6, value=round(gen_pct(0.82, 0.08), 2))
        ws.cell(row=r, column=7, value=random.choice(['↑', '→', '↓', '↑']))
    format_data_rows(ws, 2, 2 + len(SUPPLIERS) - 1, len(sc_headers))
    auto_width(ws, len(sc_headers))

    fname = f'QBR_{year}_Q{quarter}_季度報告.xlsx'
    wb.save(os.path.join(OUTPUT_DIR, fname))
    return fname

# ── Risk Report Generator ────────────────────────────────────
def generate_risk_report(year, month):
    wb = Workbook()
    period = f'{year}年{month:02d}月'

    ws = wb.active
    ws.title = 'Risk Overview'
    ws.sheet_properties.tabColor = 'C00000'
    ws['B2'] = f'風險總覽 Risk Overview — {period}'
    ws['B2'].font = TITLE_FONT
    risk_summary = [
        f'- 本月新增風險事件 {random.randint(3, 12)} 件，已結案 {random.randint(2, 8)} 件。',
        f'- 最高風險：{random.choice(RISK_TYPES)}（影響 {random.choice(PRODUCTS)}）',
        f'- 供應商風險指數：{gen_pct(0.25, 0.1):.0%}（{"正常" if random.random() > 0.3 else "偏高"}）',
        f'- 本月 P90 延遲天數：{random.randint(2, 10)} 天',
    ]
    for i, line in enumerate(risk_summary):
        ws[f'B{4+i}'] = line
        ws[f'B{4+i}'].font = BODY_FONT

    # Exception Log
    ws2 = wb.create_sheet('異常記錄 Exception Log')
    ws2.sheet_properties.tabColor = 'FFC000'
    exc_headers = ['Exception ID', 'Date', 'Category', 'Product', 'Supplier', 'Description', 'Severity', 'Status', 'Resolution']
    for c, h in enumerate(exc_headers, 1):
        ws2.cell(row=1, column=c, value=h)
    format_header_row(ws2, 1, len(exc_headers))
    num_exc = random.randint(5, 15)
    for i in range(num_exc):
        r = 2 + i
        ws2.cell(row=r, column=1, value=f'EXC-{year}{month:02d}-{i+1:03d}')
        ws2.cell(row=r, column=2, value=f'{year}/{month:02d}/{random.randint(1,28):02d}')
        ws2.cell(row=r, column=3, value=random.choice(ISSUE_CATEGORIES))
        ws2.cell(row=r, column=4, value=random.choice(PRODUCTS))
        ws2.cell(row=r, column=5, value=random.choice(SUPPLIERS))
        ws2.cell(row=r, column=6, value=random.choice([
            '交期延遲超過 SLA', '進料品質不合格', '數量短缺',
            '包裝破損', '文件不齊全', '報關延誤',
        ]))
        ws2.cell(row=r, column=7, value=random.choice(['Critical', 'High', 'Medium', 'Low']))
        ws2.cell(row=r, column=8, value=random.choice(['Open', 'Investigating', 'Resolved', 'Closed']))
        ws2.cell(row=r, column=9, value=random.choice(['補貨中', '已退貨換新', '供應商扣款', '持續追蹤', '已關閉']))
    format_data_rows(ws2, 2, 2 + num_exc - 1, len(exc_headers))
    ws2.freeze_panes = 'A2'
    auto_width(ws2, len(exc_headers), 14)

    # Risk Matrix
    ws3 = wb.create_sheet('Risk Matrix')
    ws3['B2'] = '風險矩陣 Risk Heat Map'
    ws3['B2'].font = TITLE_FONT
    matrix_headers = ['Risk Category', 'Likelihood', 'Impact', 'Score', 'Mitigation']
    for c, h in enumerate(matrix_headers, 1):
        ws3.cell(row=4, column=c+1, value=h)
    format_header_row(ws3, 4, len(matrix_headers) + 1)
    for i, risk_type in enumerate(RISK_TYPES):
        r = 5 + i
        likelihood = random.randint(1, 5)
        impact = random.randint(1, 5)
        ws3.cell(row=r, column=2, value=risk_type)
        ws3.cell(row=r, column=3, value=likelihood)
        ws3.cell(row=r, column=4, value=impact)
        ws3.cell(row=r, column=5, value=likelihood * impact)
        ws3.cell(row=r, column=6, value=random.choice(['備援供應商', '安全庫存', '合約條款', '保險', '即時監控']))
    format_data_rows(ws3, 5, 5 + len(RISK_TYPES) - 1, len(matrix_headers) + 1)
    auto_width(ws3, len(matrix_headers) + 1, 14)

    fname = f'風險報告_{year}{month:02d}_Risk_Report.xlsx'
    wb.save(os.path.join(OUTPUT_DIR, fname))
    return fname

# ── Forecast Report Generator ────────────────────────────────
def generate_forecast(year, month):
    wb = Workbook()
    period = f'{year}年{month:02d}月'

    ws = wb.active
    ws.title = 'Forecast Summary'
    ws.sheet_properties.tabColor = 'ED7D31'
    ws['B2'] = f'需求預測報告 Demand Forecast Report — {period}'
    ws['B2'].font = TITLE_FONT

    # 3-month forecast
    fc_headers = ['Product', 'Region', f'M (Actual)', f'M+1 (Forecast)', f'M+2 (Forecast)', f'M+3 (Forecast)', 'Trend', 'Confidence']
    for c, h in enumerate(fc_headers, 1):
        ws.cell(row=4, column=c, value=h)
    format_header_row(ws, 4, len(fc_headers))
    row = 5
    for prod in PRODUCTS:
        for region in random.sample(REGIONS, 2):
            base = gen_kpi_value(5000, 1500, month)
            ws.cell(row=row, column=1, value=prod)
            ws.cell(row=row, column=2, value=region)
            ws.cell(row=row, column=3, value=int(base))
            ws.cell(row=row, column=4, value=int(base * gen_pct(1.02, 0.06)))
            ws.cell(row=row, column=5, value=int(base * gen_pct(1.04, 0.08)))
            ws.cell(row=row, column=6, value=int(base * gen_pct(1.05, 0.10)))
            ws.cell(row=row, column=7, value=random.choice(['↑ Growing', '→ Stable', '↓ Declining']))
            ws.cell(row=row, column=8, value=f'{gen_pct(0.82, 0.08):.0%}')
            row += 1
    format_data_rows(ws, 5, row - 1, len(fc_headers))
    ws.freeze_panes = 'A5'
    auto_width(ws, len(fc_headers), 14)

    # Accuracy History
    ws2 = wb.create_sheet('Accuracy History')
    acc_headers = ['Month', 'MAPE', 'Bias', 'WMAPE', 'Hit Rate (±10%)', 'Model Version']
    for c, h in enumerate(acc_headers, 1):
        ws2.cell(row=1, column=c, value=h)
    format_header_row(ws2, 1, len(acc_headers))
    for i in range(6):
        m = month - 5 + i
        y = year
        if m <= 0:
            m += 12
            y -= 1
        r = 2 + i
        ws2.cell(row=r, column=1, value=f'{y}/{m:02d}')
        ws2.cell(row=r, column=2, value=round(gen_pct(0.12, 0.03), 3))
        ws2.cell(row=r, column=3, value=round(random.gauss(0.01, 0.03), 3))
        ws2.cell(row=r, column=4, value=round(gen_pct(0.11, 0.03), 3))
        ws2.cell(row=r, column=5, value=round(gen_pct(0.75, 0.08), 3))
        ws2.cell(row=r, column=6, value=f'v{random.randint(3,7)}.{random.randint(0,9)}')
    format_data_rows(ws2, 2, 7, len(acc_headers))
    auto_width(ws2, len(acc_headers))

    # Demand Drivers
    ws3 = wb.create_sheet('Demand Drivers')
    ws3['B2'] = '需求驅動因子 Demand Drivers'
    ws3['B2'].font = TITLE_FONT
    drivers = [
        f'- 季節性因素：Q{(month-1)//3+1} 為{"旺季" if month in [9,10,11,12] else "淡季"}',
        f'- 新產品 {random.choice(PRODUCTS)} 上市拉動需求 +{random.randint(5,20)}%',
        f'- {random.choice(REGIONS)} 市場{"成長" if random.random() > 0.4 else "放緩"}',
        f'- 原物料價格{"上漲" if random.random() > 0.5 else "持平"}，影響客戶備貨意願',
        f'- 競爭對手動態：{"新品發布壓力" if random.random() > 0.5 else "市場份額穩定"}',
    ]
    for i, d in enumerate(drivers):
        ws3[f'B{4+i}'] = d
        ws3[f'B{4+i}'].font = BODY_FONT
    auto_width(ws3, 2, 80)

    fname = f'需求預測_{year}{month:02d}_Demand_Forecast.xlsx'
    wb.save(os.path.join(OUTPUT_DIR, fname))
    return fname


# ── Main: Generate Full Year ─────────────────────────────────
def main():
    files = []
    print('Generating reports for 2025/03 ~ 2026/03...\n')

    for year in [2025, 2026]:
        months = range(3, 13) if year == 2025 else range(1, 4)
        for month in months:
            # Monthly MBR
            f = generate_mbr(year, month)
            files.append(f)
            print(f'  [MBR]      {f}')

            # Weekly Ops (4 weeks per month)
            for week in range(1, 5):
                f = generate_weekly_ops(year, month, week)
                files.append(f)

            # Risk Report (every month)
            f = generate_risk_report(year, month)
            files.append(f)
            print(f'  [Risk]     {f}')

            # Forecast (every month)
            f = generate_forecast(year, month)
            files.append(f)
            print(f'  [Forecast] {f}')

    # QBR (quarterly)
    for q in [2, 3, 4]:  # 2025 Q2, Q3, Q4
        f = generate_qbr(2025, q)
        files.append(f)
        print(f'  [QBR]      {f}')

    f = generate_qbr(2026, 1)  # 2026 Q1
    files.append(f)
    print(f'  [QBR]      {f}')

    print(f'\n=== Generated {len(files)} files in {OUTPUT_DIR} ===')
    print(f'  - MBR:      13 files')
    print(f'  - Weekly:   {13*4} files')
    print(f'  - Risk:     13 files')
    print(f'  - Forecast: 13 files')
    print(f'  - QBR:      4 files')

if __name__ == '__main__':
    main()
