"""
Tests for the /export-workbook endpoint (excel_export.py).

Uses FastAPI TestClient with mocked artifact loading (no real DB needed).
Verifies:
  - Response content-type is xlsx
  - Required sheet names exist
  - AI-disabled path still produces a workbook
  - Missing artifacts path produces No_Data / Export_Notes entries
  - Frontend-fallback mode works when DB is unavailable
  - Helper functions (_extract_json_block, find_download, etc.)
"""

import io
import json
import os
import sys
import unittest
from unittest.mock import patch, MagicMock

# Ensure src/ is on the path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from fastapi.testclient import TestClient
from openpyxl import load_workbook

# Import only the router and helpers (not the full app, to avoid heavy deps)
from ml.api.excel_export import (
    excel_export_router,
    _normalize_kpis,
    _pick_focus_series,
    _extract_json_block,
    find_download,
    safe_float,
    parse_csv_string,
)
from fastapi import FastAPI

# Build a minimal test app with just the export router
_test_app = FastAPI()
_test_app.include_router(excel_export_router)
client = TestClient(_test_app)

FAKE_DB_URL = "postgresql://test:test@localhost:5432/testdb"

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

MINIMAL_ARTIFACTS = {
    "replay_metrics": {
        "with_plan": {
            "service_level": 0.95,
            "stockout_units": 120,
            "holding_units": 5000,
            "total_cost": 45000,
        },
        "without_plan": {
            "service_level": 0.82,
            "stockout_units": 800,
            "holding_units": 3000,
            "total_cost": 62000,
        },
    },
    "forecast_series": {
        "groups": [
            {
                "key": "MAT-001__PLANT-A",
                "material_code": "MAT-001",
                "plant_id": "PLANT-A",
                "points": [
                    {"date": "2026-01-01", "actual": 100, "p50": 95, "p90": 110},
                    {"date": "2026-01-08", "actual": 120, "p50": 115, "p90": 130},
                    {"date": "2026-01-15", "actual": 90, "p50": 100, "p90": 120},
                ],
            }
        ]
    },
    "plan_csv": "material_code,plant_id,order_qty,order_date\nMAT-001,PLANT-A,500,2026-02-01\nMAT-001,PLANT-A,300,2026-02-08\n",
    "inventory_projection": [
        {"period": "2026-01", "with_plan": 1000, "without_plan": 400},
        {"period": "2026-02", "with_plan": 800, "without_plan": 100},
    ],
    "report_json": {
        "summary": "Planning run completed successfully with 95% service level."
    },
    "constraint_check": {
        "checks": [
            {"name": "MOQ", "status": "passed", "details": "All MOQs met."},
            {"name": "Lead Time", "status": "failed", "details": "2 items exceed lead time."},
        ]
    },
}

MINIMAL_RUN_META = {
    "id": 42,
    "workflow": "workflow_A_replenishment",
    "status": "succeeded",
    "started_at": "2026-02-21T10:00:00",
    "finished_at": "2026-02-21T10:05:00",
    "meta": None,
}


def _mock_load_run_artifacts(run_id):
    """Return a copy of minimal artifacts with __notes__."""
    arts = {k: (json.loads(json.dumps(v)) if isinstance(v, (dict, list)) else v) for k, v in MINIMAL_ARTIFACTS.items()}
    arts["__notes__"] = []
    return arts


def _mock_load_run_artifacts_empty(run_id):
    return {"__notes__": []}


def _mock_load_run_meta(run_id):
    return dict(MINIMAL_RUN_META, run_id=run_id)


# ---------------------------------------------------------------------------
# Tests – DB-primary path
# ---------------------------------------------------------------------------

class TestExportWorkbookEndpoint(unittest.TestCase):
    """Tests for POST /export-workbook (DB-primary path)."""

    @patch("ml.api.excel_export._get_db_url", return_value=FAKE_DB_URL)
    @patch("ml.api.excel_export.load_run_meta", side_effect=_mock_load_run_meta)
    @patch("ml.api.excel_export.load_run_artifacts", side_effect=_mock_load_run_artifacts)
    @patch("ml.api.excel_export.DEEPSEEK_API_KEY", "")
    def test_basic_export_produces_valid_xlsx(self, mock_arts, mock_meta, mock_db):
        resp = client.post("/export-workbook", json={
            "version": "v1",
            "run_id": 42,
            "ai_insights": False,
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resp.headers["content-type"],
        )
        wb = load_workbook(io.BytesIO(resp.content))
        sheet_names = wb.sheetnames
        self.assertIn("Executive_Summary", sheet_names)
        self.assertIn("KPI_Dashboard", sheet_names)
        self.assertIn("Forecast_Chart", sheet_names)
        self.assertIn("Plan_Output", sheet_names)
        wb.close()

    @patch("ml.api.excel_export._get_db_url", return_value=FAKE_DB_URL)
    @patch("ml.api.excel_export.load_run_meta", side_effect=_mock_load_run_meta)
    @patch("ml.api.excel_export.load_run_artifacts", side_effect=_mock_load_run_artifacts)
    @patch("ml.api.excel_export.DEEPSEEK_API_KEY", "")
    def test_ai_disabled_still_produces_workbook(self, mock_arts, mock_meta, mock_db):
        resp = client.post("/export-workbook", json={
            "version": "v1",
            "run_id": 42,
            "ai_insights": True,  # requested but key is empty
        })
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertIn("Executive_Summary", wb.sheetnames)
        if "Export_Notes" in wb.sheetnames:
            notes_ws = wb["Export_Notes"]
            notes_text = " ".join(
                str(cell.value or "") for row in notes_ws.iter_rows() for cell in row
            )
            self.assertIn("AI", notes_text)
        wb.close()

    @patch("ml.api.excel_export._get_db_url", return_value=FAKE_DB_URL)
    @patch("ml.api.excel_export.load_run_meta", side_effect=_mock_load_run_meta)
    @patch("ml.api.excel_export.load_run_artifacts", side_effect=_mock_load_run_artifacts_empty)
    @patch("ml.api.excel_export.DEEPSEEK_API_KEY", "")
    def test_missing_artifacts_produces_no_data_sheet(self, mock_arts, mock_meta, mock_db):
        resp = client.post("/export-workbook", json={
            "version": "v1",
            "run_id": 999,
            "ai_insights": False,
        })
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertIn("No_Data", wb.sheetnames)
        wb.close()

    @patch("ml.api.excel_export._get_db_url", return_value=FAKE_DB_URL)
    @patch("ml.api.excel_export.load_run_meta", side_effect=_mock_load_run_meta)
    @patch("ml.api.excel_export.load_run_artifacts", side_effect=_mock_load_run_artifacts)
    @patch("ml.api.excel_export.DEEPSEEK_API_KEY", "")
    def test_focus_series_key_selects_correct_series(self, mock_arts, mock_meta, mock_db):
        resp = client.post("/export-workbook", json={
            "version": "v1",
            "run_id": 42,
            "ai_insights": False,
            "focus": {
                "series_key": "MAT-001__PLANT-A",
                "sku": "MAT-001",
                "plant": "PLANT-A",
                "mode": "selected",
            },
        })
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertIn("Forecast_Chart", wb.sheetnames)
        fc = wb["Forecast_Chart"]
        self.assertGreaterEqual(fc.max_row, 4)
        wb.close()

    @patch("ml.api.excel_export._get_db_url", return_value=FAKE_DB_URL)
    @patch("ml.api.excel_export.load_run_meta", side_effect=_mock_load_run_meta)
    @patch("ml.api.excel_export.load_run_artifacts", side_effect=_mock_load_run_artifacts)
    @patch("ml.api.excel_export.DEEPSEEK_API_KEY", "")
    def test_content_disposition_includes_run_id(self, mock_arts, mock_meta, mock_db):
        resp = client.post("/export-workbook", json={
            "version": "v1",
            "run_id": 42,
            "ai_insights": False,
        })
        cd = resp.headers.get("content-disposition", "")
        self.assertIn("42", cd)
        self.assertIn(".xlsx", cd)

    @patch("ml.api.excel_export._get_db_url", return_value=FAKE_DB_URL)
    @patch("ml.api.excel_export.load_run_meta", side_effect=_mock_load_run_meta)
    @patch("ml.api.excel_export.load_run_artifacts", side_effect=_mock_load_run_artifacts)
    @patch("ml.api.excel_export.DEEPSEEK_API_KEY", "test-key")
    @patch("ml.api.excel_export.generate_ai_insights", return_value={
        "executive_summary": "Test summary.",
        "key_findings": ["Finding 1"],
        "recommendations": ["Rec 1"],
        "risk_assessment": "Low risk.",
    })
    def test_ai_insights_included_when_available(self, mock_ai, mock_arts, mock_meta, mock_db):
        resp = client.post("/export-workbook", json={
            "version": "v1",
            "run_id": 42,
            "ai_insights": True,
        })
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        es = wb["Executive_Summary"]
        all_text = " ".join(
            str(cell.value or "") for row in es.iter_rows() for cell in row
        )
        self.assertIn("Test summary", all_text)
        wb.close()

    @patch("ml.api.excel_export._get_db_url", return_value=FAKE_DB_URL)
    @patch("ml.api.excel_export.load_run_meta", side_effect=_mock_load_run_meta)
    @patch("ml.api.excel_export.load_run_artifacts", side_effect=_mock_load_run_artifacts)
    @patch("ml.api.excel_export.DEEPSEEK_API_KEY", "")
    def test_how_to_read_section_present(self, mock_arts, mock_meta, mock_db):
        """Executive Summary should contain the How to Read guide."""
        resp = client.post("/export-workbook", json={
            "version": "v1",
            "run_id": 42,
            "ai_insights": False,
        })
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        es = wb["Executive_Summary"]
        all_text = " ".join(
            str(cell.value or "") for row in es.iter_rows() for cell in row
        )
        self.assertIn("How to Read This Workbook", all_text)
        self.assertIn("KPI_Dashboard", all_text)
        wb.close()

    @patch("ml.api.excel_export._get_db_url", return_value=FAKE_DB_URL)
    @patch("ml.api.excel_export.load_run_meta", side_effect=_mock_load_run_meta)
    @patch("ml.api.excel_export.load_run_artifacts", side_effect=_mock_load_run_artifacts)
    @patch("ml.api.excel_export.DEEPSEEK_API_KEY", "")
    def test_risk_analysis_sheet_always_created(self, mock_arts, mock_meta, mock_db):
        """Risk_Analysis sheet should exist even when no risk data."""
        resp = client.post("/export-workbook", json={
            "version": "v1",
            "run_id": 42,
            "ai_insights": False,
        })
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertIn("Risk_Analysis", wb.sheetnames)
        wb.close()


# ---------------------------------------------------------------------------
# Tests – Frontend-fallback path (no DB)
# ---------------------------------------------------------------------------

class TestFrontendFallback(unittest.TestCase):
    """Tests for the frontend-data fallback path (no DB URL configured)."""

    @patch("ml.api.excel_export._get_db_url", return_value=None)
    @patch("ml.api.excel_export.DEEPSEEK_API_KEY", "")
    def test_frontend_data_produces_workbook(self, mock_db):
        """When no DB URL is set, frontend-provided data should be used."""
        resp = client.post("/export-workbook", json={
            "version": "v1",
            "run_id": 42,
            "ai_insights": False,
            "run_meta": {"run_id": 42, "status": "succeeded", "workflow": "workflow_A"},
            "chart_payload": {
                "inventory_projection": [
                    {"period": "2026-01", "with_plan": 500, "without_plan": 200},
                ],
            },
            "downloads": [
                {
                    "label": "replay_metrics",
                    "fileName": "replay_metrics.json",
                    "content": json.dumps({
                        "with_plan": {"service_level": 0.90},
                        "without_plan": {"service_level": 0.75},
                    }),
                    "mimeType": "application/json",
                },
                {
                    "label": "plan.csv",
                    "fileName": "plan.csv",
                    "content": "material_code,plant_id,order_qty\nX,P1,100\n",
                    "mimeType": "text/csv",
                },
            ],
        })
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertIn("Executive_Summary", wb.sheetnames)
        self.assertIn("Plan_Output", wb.sheetnames)
        # Export_Notes should mention frontend fallback
        if "Export_Notes" in wb.sheetnames:
            notes_text = " ".join(
                str(cell.value or "") for row in wb["Export_Notes"].iter_rows() for cell in row
            )
            self.assertIn("frontend", notes_text.lower())
        wb.close()

    @patch("ml.api.excel_export._get_db_url", return_value=None)
    @patch("ml.api.excel_export.DEEPSEEK_API_KEY", "")
    def test_no_data_at_all_produces_no_data_sheet(self, mock_db):
        """When neither DB nor frontend data is available, return No_Data."""
        resp = client.post("/export-workbook", json={
            "version": "v1",
            "ai_insights": False,
        })
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertIn("No_Data", wb.sheetnames)
        wb.close()

    @patch("ml.api.excel_export._get_db_url", return_value=FAKE_DB_URL)
    @patch("ml.api.excel_export.load_run_meta", side_effect=Exception("DB down"))
    @patch("ml.api.excel_export.load_run_artifacts", side_effect=Exception("DB down"))
    @patch("ml.api.excel_export.DEEPSEEK_API_KEY", "")
    def test_db_failure_falls_back_to_frontend(self, mock_arts, mock_meta, mock_db):
        """When DB fails, should fall back to frontend data."""
        resp = client.post("/export-workbook", json={
            "version": "v1",
            "run_id": 42,
            "ai_insights": False,
            "run_meta": {"run_id": 42, "status": "succeeded"},
            "downloads": [
                {
                    "label": "plan.csv",
                    "fileName": "plan.csv",
                    "content": "material_code,order_qty\nA,100\n",
                    "mimeType": "text/csv",
                },
            ],
        })
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertIn("Executive_Summary", wb.sheetnames)
        self.assertIn("Plan_Output", wb.sheetnames)
        wb.close()


# ---------------------------------------------------------------------------
# Tests – Helper functions
# ---------------------------------------------------------------------------

class TestHelpers(unittest.TestCase):
    """Tests for data helper functions."""

    def test_safe_float_numbers(self):
        self.assertEqual(safe_float(42), 42.0)
        self.assertEqual(safe_float("3.14"), 3.14)
        self.assertIsNone(safe_float("abc"))
        self.assertIsNone(safe_float(None))
        self.assertIsNone(safe_float(float("nan")))

    def test_parse_csv_string(self):
        csv_str = "a,b,c\n1,2,3\n4,5,6"
        rows = parse_csv_string(csv_str)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["a"], "1")
        self.assertEqual(rows[1]["c"], "6")

    def test_parse_csv_empty(self):
        self.assertEqual(parse_csv_string(""), [])
        self.assertEqual(parse_csv_string(None), [])

    def test_normalize_kpis_standard(self):
        raw = {
            "with_plan": {"service_level": 0.95, "stockout_units": 100},
            "without_plan": {"service_level": 0.80, "stockout_units": 500},
        }
        result = _normalize_kpis(raw)
        self.assertIn("delta", result)
        self.assertAlmostEqual(result["delta"]["service_level"], 0.15)
        self.assertAlmostEqual(result["delta"]["stockout_units"], -400)

    def test_normalize_kpis_flat(self):
        raw = {"service_level": 0.9, "total_cost": 1000}
        result = _normalize_kpis(raw)
        self.assertEqual(result["with_plan"]["service_level"], 0.9)

    def test_normalize_kpis_none(self):
        self.assertIsNone(_normalize_kpis(None))
        self.assertIsNone(_normalize_kpis("not a dict"))

    def test_pick_focus_series_by_key(self):
        groups = [
            {"key": "A", "material_code": "MAT-A", "plant_id": "P1", "points": [{"x": 1}]},
            {"key": "B", "material_code": "MAT-B", "plant_id": "P2", "points": [{"x": 2}]},
        ]
        from ml.api.excel_export import FocusSpec
        focus = FocusSpec(series_key="B")
        result = _pick_focus_series(groups, focus)
        self.assertEqual(result["key"], "B")

    def test_pick_focus_series_by_sku(self):
        groups = [
            {"key": "A", "material_code": "MAT-A", "plant_id": "P1", "points": []},
            {"key": "B", "material_code": "MAT-B", "plant_id": "P2", "points": []},
        ]
        from ml.api.excel_export import FocusSpec
        focus = FocusSpec(sku="mat-b")  # case-insensitive
        result = _pick_focus_series(groups, focus)
        self.assertEqual(result["key"], "B")

    def test_pick_focus_series_fallback_first(self):
        groups = [{"key": "X", "points": []}]
        result = _pick_focus_series(groups, None)
        self.assertEqual(result["key"], "X")

    def test_pick_focus_series_empty(self):
        self.assertIsNone(_pick_focus_series([], None))

    # ---- _extract_json_block ----

    def test_extract_json_block_direct(self):
        result = _extract_json_block('{"key": "value"}')
        self.assertEqual(result, {"key": "value"})

    def test_extract_json_block_with_fences(self):
        result = _extract_json_block('```json\n{"key": "value"}\n```')
        self.assertEqual(result, {"key": "value"})

    def test_extract_json_block_embedded(self):
        text = 'Here is the result:\n{"a": 1, "b": 2}\nEnd.'
        result = _extract_json_block(text)
        self.assertEqual(result, {"a": 1, "b": 2})

    def test_extract_json_block_invalid(self):
        self.assertIsNone(_extract_json_block("no json here"))

    # ---- find_download ----

    def test_find_download_match(self):
        dls = [
            {"label": "report.json", "content": "{}"},
            {"label": "plan.csv", "content": "a,b\n1,2"},
        ]
        result = find_download(dls, "plan")
        self.assertEqual(result["label"], "plan.csv")

    def test_find_download_no_match(self):
        dls = [{"label": "foo", "content": "x"}]
        self.assertIsNone(find_download(dls, "bar"))

    def test_find_download_empty(self):
        self.assertIsNone(find_download([], "x"))
        self.assertIsNone(find_download(None, "x"))

    def test_find_download_case_insensitive(self):
        dls = [{"label": "Replay_Metrics.json", "content": "{}"}]
        result = find_download(dls, "replay_metrics")
        self.assertIsNotNone(result)


if __name__ == "__main__":
    unittest.main()
